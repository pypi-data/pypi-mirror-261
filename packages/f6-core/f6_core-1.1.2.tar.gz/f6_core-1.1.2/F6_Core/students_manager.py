# Copyright 2024 Degtyarev Ivan

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#     http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS
import json
import os, io
import datetime
import calendar

from . encryption_data import EncryptionData


class Student:

    def __init__(self, fio: str, sick_days=None, absence_days=None):
        if not sick_days:
            sick_days = {}
        if not absence_days:
            absence_days = {}
        self.__fio = self.check_fio(fio)
        self.__sick_days = self.check_days(sick_days)
        self.__absence_days = self.check_days(absence_days)
        self.marks = {}

    def __str__(self):
        return self.create_shorts_fio(self.fio)

    @classmethod
    def check_days(cls, days: dict) -> dict:
        """Проверяет правильность словаря дней, при верном результате возвращает его, иначе ошибку ValueError"""
        if not isinstance(days, dict):
            raise ValueError('Функция работает только с данными типа словарь')
        for day, hours in days.items():
            if not (cls.is_valud_day(day) and cls.is_valud_hours(hours)):
                raise ValueError('Словарь с днями поврежден')
        return days

    @classmethod
    def check_fio(cls, fio: str) -> str:
        """Проверяет правильность ФИО, при верном результате возвращает ФИО, иначе ошибку ValueError"""
        if cls.is_valud_fio(fio):
            return ' '.join(fio.title().split())
        raise ValueError('ФИО должно состоять только из букв и быть из 3 частей, каждая из которых не менее 2 символов')

    @staticmethod
    def is_valud_fio(fio: str) -> bool:
        '''Проверяет правильность ФИО, возвращает булевое значение'''
        return isinstance(fio.replace(' ', ''), str) and sum(
            [item.isalpha() and len(item) >= 2 for item in fio.split()]) == 3

    @classmethod
    def check_day(cls, day: int) -> int:
        '''Проверяет правильность дня, при верном результате возвращает день, иначе ошибку ValueError'''
        if cls.is_valud_day(day):
            return day
        raise ValueError('День должен быть целым числом, от 1 до 31')

    @staticmethod
    def is_valud_day(day: int) -> bool:
        """Проверяет правильность дня, возвращает булевое значение"""
        return isinstance(day, int) and 1 <= day <= 31

    @classmethod
    def check_hours(cls, hours: int) -> int:
        '''Проверяет правильность "часов", при верном результате возвращает "часы", иначе ошибку ValueError'''
        if cls.is_valud_hours(hours):
            return hours
        raise ValueError('Часы должены быть записаны целым числом, от 0 до 10')

    @staticmethod
    def is_valud_hours(hours: int) -> bool:
        """Проверяет правильность часов, возвращает булевое значение"""
        return isinstance(hours, int) and 0 <= hours <= 10

    @property
    def fio(self) -> str:
        return self.__fio

    @fio.setter
    def fio(self, value: str) -> None:
        self.__fio = self.check_fio(value)

    @classmethod
    def create_shorts_fio(cls, fio: str) -> str:
        '''Создает сокращенную форму ФИО: ФФФФФФФ И. О., возвращает строку'''

        fio = cls.check_fio(fio)

        last_name, first_name, middle_name = fio.split()
        return f'{last_name} {first_name[0]}. {middle_name[0]}.'

    @property
    def sick_days(self) -> list:
        return self.__sick_days

    @property
    def absence_days(self) -> list:
        return self.__absence_days

    def add_sick_day(self, day: int, hours: int) -> None:
        """Добавляет пару: ключ - день и значение - время в словарь с прогулами по болезни"""
        day = self.check_day(day)
        hours = self.check_hours(hours)
        self.__sick_days[day] = hours

    def add_absence_day(self, day: int, hours: int) -> None:
        """Добавляет пару: ключ - день и значение - время в словарь с прогулами по неуважительной причине"""
        day = self.check_day(day)
        hours = self.check_hours(hours)
        self.__absence_days[day] = hours

    def get_statistic_for_student(self) -> dict:
        """
        Возвращает статистику по студенту в виде словаря
        {
        FIO: str,
        Sick_days: (list, int),
        Absence_days: (list, int),
        }
         """
        return {
            'FIO': self.fio,
            "Sick_days": (self.sick_days, sum(self.sick_days.values())),
            "Absence_days": (self.absence_days, sum(self.absence_days.values())),
        }

    def add_day(self, number_day: int, hours: int, type_day='a') -> None:
        """
        Добавляет номер дня и его значения в соответсвующий словарь в зависимости от типа:
        type_day='a'/'s'
        """
        type_day = type_day.lower()
        if type_day == 's':
            self.add_sick_day(number_day, hours)
        elif type_day == 'a':
            self.add_absence_day(number_day, hours)
        else:
            raise ValueError("Неверно указан тип дня")


class ManagerStudents:
    ENCRYPTION_CLASS = EncryptionData
    CLASS_STUDENT = Student
    MONTHS = (
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    )

    def __init__(self, period: tuple, user=None, days=None, students=None):
        self.user = user
        self.period = period  # (month, year)
        self.days = days if days else self.generate_work_days(month=self.period[0], year=self.period[1],
                                                              happy_days=user.happy_days)
        self.__students = [] if not students else students
        self.date = datetime.datetime.now().timetuple()
        self.parametrs = {}
        self.couples = {}
        self.is_archive = False

    @classmethod
    def generate_work_days(cls, month: int, year: int, happy_days=None) -> dict:
        """
        Генерирует рабочие дни с учетом праздников и выходных
        Возвращает словарь в виде
        {
        int: int,
        "day": "hours"
        }
        """
        if not happy_days:
            happy_days = {}
        month = cls.check_month(month)
        year = cls.check_year(year)
        work_days = {}
        for week in calendar.monthcalendar(month=month, year=year):
            for day in range(7):
                if day > 4 or week[day] in happy_days.get(str(month), []):
                    continue
                elif week[day] != 0 and week[day]:
                    work_days[week[day]] = 0
        return work_days

    @classmethod
    def check_month(cls, month: int) -> int:
        if cls.is_valud_month(month):
            return month
        raise ValueError('Месяц должен быть от 1 до 12')

    @staticmethod
    def is_valud_month(month: int) -> bool:
        return isinstance(month, int) and 1 <= month <= 12

    @classmethod
    def check_year(cls, year: int) -> int:
        if cls.is_valud_year(year):
            return year
        raise ValueError('Год должен быть от 1700')

    @staticmethod
    def is_valud_year(year: int) -> bool:
        return isinstance(year, int) and 1700 <= year

    def off_day(self, day: int, is_happy_day=False, month=0) -> None:
        """Удаляет рабочий день из переменной self.days"""
        self.CLASS_STUDENT.check_day(day)
        if not month or month == self.period[0]:
            if day in self.days:
                for st in self.students:
                    if st.sick_days.get(day):
                        del st.sick_days[day]
                    elif st.absence_days.get(day):
                        del st.absence_days[day]
                del self.days[day]
        if is_happy_day:
            self.user.happy_days[str(self.period[0] if not month else month)] = self.user.happy_days.get(str(self.period[0] if not month else month), set()) | {day}

    def on_day(self, day: int, hours: int = 0, is_happy_day=False, month=0) -> None:
        """Добавляет рабочий день из переменной self.days"""
        self.CLASS_STUDENT.check_day(day)
        self.CLASS_STUDENT.check_hours(hours)
        if not month or month == self.period[0]:
            self.days[day] = hours

        if is_happy_day:
            if day in self.user.happy_days.get(str(self.period[0] if not month else month)):
                self.user.happy_days[str(self.period[0] if not month else month)].remove(day)

    def add_day(self, student, number_day, hours, type_day='a'):
        type_day = type_day.lower()
        if type_day == 's':
            student.add_day(number_day, hours, type_day)
        elif type_day == 'a':
            student.add_day(number_day, hours, type_day)
        else:
            raise ValueError("Неверно указан тип дня")

    def del_day(self, student, number_day):
        if number_day in student.sick_days:
            del student.sick_days[number_day]
        if number_day in student.absence_days:
            del student.absence_days[number_day]

    @staticmethod
    def is_valud_period(period: tuple) -> bool:
        return isinstance(period, tuple) and len(period) == 2 and \
            isinstance(period[0], int) and isinstance(period[1], int) and 1 <= period[0] <= 12 and 1700 <= period[1]

    @classmethod
    def check_period(cls, period: tuple) -> tuple:
        if cls.is_valud_period(period):
            return period
        raise ValueError(
            'Период должен быть катрежем из двух чисел: месяц (int) и год (int), которые находятся в диапазоне от 1 до 12 (month) и от 1700 (year) соответственно ')

    def set_period(self, month: int, year: int):
        '''Изменяет месяц и год таблицы'''
        self.period = self.check_period((month, year))
        self.days = self.generate_work_days(month, year, self.user.happy_days)

    def add_hours_by_day(self, day, hours: int = 0) -> None:
        """Добавляет значение времени в словарь по ключу day"""
        if day in self.days and 0 < hours <= 10:
            self.days[day] = hours

    def __getitem__(self, indx: int) -> Student:
        """Врзвращает эксезмпляр студента из списка студентов"""
        return self.students[indx]

    def __setitem__(self, indx: int, student: Student) -> None:
        """Изменяет экземпляр студента по его индеку с списке"""
        self.__students[indx] = student

    def __delitem__(self, key: int) -> None:
        self.remove_student_id(key)

    @property
    def students(self) -> list:
        return self.__students

    def add_student(self, student: Student) -> None:
        self.__students.append(student)

    def remove_student_id(self, id_s: int) -> None:
        del self.__students[id_s]

    def remove_student_obj(self, obj: Student) -> None:
        self.__students.remove(obj)

    def save_students(self, file_name='students.json', path_file=None) -> None:
        """
        Сохраняет данные о студентах
        """
        self.ENCRYPTION_CLASS.SECURITY_LEVEL = self.user.security_level
        parser = self.ENCRYPTION_CLASS(self.user.password, self.user.user_id)
        
        if path_file is None:
            path_file = self.user.path
        date = self.__serialization()

        if not os.path.isdir(self.user.path):
            os.mkdir(self.user.path)

        if self.user.security_level == 0 or self.user.security_level == 1:
            liststudents, psc, R = parser.encode(date['Liststudents'])
            date['Liststudents'] = liststudents
            date['PSC'] = psc
            date['Rang'] = R
            
            with open(os.path.join(path_file, file_name), 'w') as f:
                json.dump(date, f)
        else:
            date = json.dumps(date).encode('UTF-16')
            sequence_byte, psc, R = parser.encode(date)
            with open(os.path.join(path_file, file_name), 'wb') as f:
                f.write(sequence_byte.getvalue())
            

    def load_students(self, file_name='students.json', user_path=None) -> dict | bytes:
        """Читает данные из файла и возвращает из них словарь"""
        if user_path is None:
            user_path = self.user.path

        if self.user.security_level == 0 or self.user.security_level == 1:
            with open(os.path.join(user_path, file_name), 'r') as f:
                data = json.load(f)
            return data
        else:
            with open(os.path.join(user_path, file_name), 'rb') as f:
                file_content = io.BytesIO(f.read())

            return file_content

    def __serialization(self) -> dict:
        """Преобразует данные для сохранения в Json"""
        liststudents, psc, R = self.ENCRYPTION_CLASS.convert_list_to_str(self.__convert()), '', ''

        data = {
            "Period": self.period,
            "Liststudents": liststudents,
            "Couples": self.couples,
            'Days': self.days,
            "Parametrs": self.parametrs,
            'PSC': '',
            'Rang': '',
        }

        return data

    @staticmethod
    def crate_eternal_iter(data: str, indx: int = 0) -> str:
        """Создает из итерируемого объекта бесконечный итератор"""
        while True:
            yield data[indx]
            indx += 1
            if indx == len(data):
                indx = 0

    def __convert(self) -> list:
        result = []
        for st in self.__students:
            result.extend(
                [st.fio, ''.join([str(i).rjust(2, '0') + str(st.sick_days[i]).rjust(2, '0') for i in st.sick_days]),
                 ''.join([str(i).rjust(2, '0') + str(st.absence_days[i]).rjust(2, '0') for i in st.absence_days]),
                 ''.join([str(k).rjust(2, '0') + str(st.marks[k]).rjust(2, '0') for k in st.marks])])
        return result

    @staticmethod
    def convert_str_to_list(date: str, sep: str = '!'):
        """Преобразует данные из строки в экземпляры класса Student"""

        def convert_data(data: str):
            result = {}
            for i in range(len(data) // 4):
                day = int(data[i * 4] + data[i * 4 + 1])
                hours = int(data[i * 4 + 2] + data[i * 4 + 3])

                result[day] = hours
            return result

        date = date[2:-2].split(sep)
        students = []
        for i in range(len(date) // 4):
            fio = date[i * 4]
            s_d = date[i * 4 + 1]
            a_d = date[i * 4 + 2]
            marks = date[i * 4 + 3]
            s = Student(fio, convert_data(s_d), convert_data(a_d))

            result = {}
            for i in range(0, len(marks), 4):
                raw = marks[i:i + 4]
                result[int(raw[0:2])] = int(raw[2:])
            s.marks = result
            students.append(s)

        return students

    @classmethod
    def load_manager_students(cls, user, data=None, file_name='students.json'):

        if data is None:
            user_path = user.path
            if user.security_level == 1 or user.security_level == 0:
                with open(os.path.join(user_path, file_name), 'r') as f:
                    data = json.load(f)
            else:
                with open(os.path.join(user_path, file_name), 'rb') as f:
                    data = io.BytesIO(f.read())

        cls.ENCRYPTION_CLASS.SECURITY_LEVEL = user.security_level
        parser = cls.ENCRYPTION_CLASS(user.password, user.user_id)

        if user.security_level == 0 or user.security_level == 1:
            liststudents = parser.decode(data['Liststudents'], data['PSC'])
        else:
            data = parser.decode(data)
            liststudents = data['Liststudents']

        new_obj = ManagerStudents(data['Period'], user, {int(k): int(v) for k, v in data['Days'].items()})


        students = cls.convert_str_to_list(liststudents)
        students = list(sorted(students, key=lambda item: item.fio))
        for student in students:
            new_obj.add_student(student)

        new_obj.parametrs = data['Parametrs']
        new_obj.couples = data.get('Couples', {})
        return new_obj

    def get_statistics(self):
        all_hours = sum(list(self.days.values()))
        sack_days_hours = 0
        absence_days_hours = 0
        for student in self.students:
            statistic = student.get_statistic_for_student()
            sack_days_hours += statistic['Sick_days'][1]
            absence_days_hours += statistic['Absence_days'][1]

        all_absence = sack_days_hours + absence_days_hours
        man_hours = len(self.students) * all_hours
        statisitcs = {
            'all_hours': all_hours,
            'sack_days_hours': sack_days_hours,
            'absence_days_hours': absence_days_hours,
            'all_absence': all_absence,
            'man_hours': man_hours,
        }
        if man_hours > 0:
            total_attendance = (man_hours - (absence_days_hours + sack_days_hours)) / man_hours
            quality_attendance = (man_hours - absence_days_hours) / man_hours
            absences_by_student = absence_days_hours / len(self.students)

            statisitcs['total_attendance'] = total_attendance
            statisitcs["quality_attendance"] = quality_attendance
            statisitcs['absences_by_student'] = absences_by_student
        else:
            statisitcs['total_attendance'] = 0
            statisitcs["quality_attendance"] = 0
            statisitcs['absences_by_student'] = 0

        return statisitcs

    def get_statistics_marks(self):
        statisitcs = {
            'heaving_2': 0,
            "heaving_5": 0,
            'heaving_one_3': 0,
            "heaving_4_and_5": 0,
            'is_ready': 0,
            'no_att': 0,
        }
        for i in self.students:
            statisitcs['is_ready'] = 1
            marks = [st for st in i.marks.values()]
            if len(marks) >= len(self.couples):
                if 2 in marks:
                    statisitcs['heaving_2'] += 1
                elif all(map(lambda x: x == 5, marks)):
                    statisitcs['heaving_5'] += 1
                elif marks.count(3) == 1:
                    statisitcs['heaving_one_3'] += 1
                else:
                    statisitcs['heaving_4_and_5'] += 1
        count_student = len(self.students)
        if count_student > 0:
            statisitcs['total_academic_performance'] = ((count_student - statisitcs['heaving_2']) / count_student)
            statisitcs['quality_academic_performance'] = (statisitcs['heaving_4_and_5'] / count_student)
        else:
            statisitcs['total_academic_performance'] = 0
            statisitcs['quality_academic_performance'] = 0
        return statisitcs

    def push_archive(self, *args, **kargs) -> None:
        '''Переносит файл в архив и переименновывает ее новым именем'''
        if not os.path.exists(os.path.join(self.user.path, 'archive')):
            os.makedirs(os.path.join(self.user.path, "archive"))
        file_path, file_name = self.creat_file_path_in_archive(*args, **kargs)

        self.replace_file(os.path.join(self.user.path, 'students.json'),
                          os.path.join(self.user.path, 'archive', 'students.json'))
        self.rename_file(os.path.join(self.user.path, 'archive', 'students.json'), file_path)

    def creat_file_path_in_archive(self, file_name=None, auto: bool = False) -> tuple:
        '''Сделает путь для текущий активный фалй students.json в архив,
        при этом если атрибут auto=False, но файл уже сеть в директории, то выдаст ошибку FileExistsError.
        Иначе будет автоматически измененно имя файла, который будет перемещен в архив.
        Возвращает: (file_path, file_name)
        '''
        if not file_name:
            file_name = self.create_archive_file_name()

        if os.path.exists(os.path.join(self.user.path, 'archive', file_name)) and not auto:
            raise FileExistsError

        id_copy = 1
        while os.path.exists(os.path.join(self.user.path, 'archive', file_name)):
            file_name = self.create_archive_file_name(id_copy)
            id_copy += 1

        return os.path.join(self.user.path, 'archive', file_name), file_name

    def create_archive_file_name(self, level: int = '') -> str:
        '''Создает имя для архивного файла'''
        if level or level == 0:
            level = ' (' + str(level) + ')'
        return str(self.period[1]) + '_' + str(self.period[0]).rjust(2, '0') + self.MONTHS[
            self.period[0] - 1] + level + '.json'

    def create_new_table(self) -> None:
        """Преобразует текущий экземпляр класса в месяц следующий по году и сохраняет новую таблицу в students.json"""
        if 1 <= self.period[0] < 12:
            month = self.period[0] + 1
            year = self.period[1]
            if month == 7 or month == 8:
                month = 9
        else:
            month = 1
            year = self.period[1] + 1

        self.set_period(month, year)
        self.couples.clear()
        for s in self.students:
            s.sick_days.clear()
            s.absence_days.clear()
            s.marks.clear()
        self.save_students()

    def replace_file(self, path1: str, path2: str) -> None:
        """Перемещает файл из каталога path1 в каталог path2"""
        os.replace(path1, path2)

    def rename_file(self, path1: str, path2: str) -> None:
        '''Переименновывает файл из path1 в path2'''
        os.rename(path1, path2)

    def init_archive(self) -> tuple:
        """
        Создает каталог для архиа, если он отсутствует.
        Возвращает путь до каталога и спиок имен файлов в каталоге.
        """
        path_archive = os.path.join(self.user.path, 'archive')
        if not os.path.exists(os.path.join(path_archive)):
            os.makedirs(os.path.join(path_archive))
        walk = tuple(os.walk(path_archive))
        self.path_archive_files = walk[0][0]
        self.name_files = sorted([i for i in walk[0][-1] if i.endswith('.json') and len(i) >= 15], reverse=True)
        return self.path_archive_files, self.name_files

    def load_archive_file(self, file_name: str, file_path=None):
        '''Возвращает экземпляр архивного месяца'''
        if file_name in self.name_files:
            return self.load_manager_students(self.user, self.load_students(file_name=file_name, user_path=file_path),
                                              file_name)
        return self

    def clear_marks(self) -> None:
        """Очищает все оценки для студентов, включая названия предменат и ФИО преподавателя"""
        for i in range(len(self.students)):
            self.students[i].marks.clear()

    def clear_absences(self) -> None:
        """Очищает все прогулы для студентов"""
        for i in range(len(self.students)):
            self.students[i].sick_days.clear()
            self.students[i].absence_days.clear()

    def get_statistics_for_group(self) -> dict:
        statistics_resalt = {}

        for i in self.students:
            statistics = i.get_statistic_for_student()
            resalt = statistics['Sick_days'][1] + statistics['Absence_days'][1]
            if resalt > 0:
                statistics_resalt[statistics['FIO']] = resalt

        statistics_resalt = dict(sorted(statistics_resalt.items(), key=lambda item: item[1], reverse=True))
        return statistics_resalt

    def get_total_statistic_period(self) -> dict:
        """
        Генерирует статистику за все месяцы в виде словаря
        {
        tuple: tuple,
        (int, int): (int, int, int),
        (month, year): (посещ.качест, посещ.общая, всего прогулов)
        }
        """
        months = {}
        path_archive_files, name_files = self.init_archive()
        for i in name_files:
            month = self.load_archive_file(i, path_archive_files)
            statistic = month.get_statistics()
            months[(month.period[1], month.period[0] - 1)] = (
            statistic['sack_days_hours'], statistic['absence_days_hours'], statistic['all_absence'])

        months = dict(sorted(months.items(), key=lambda item: item))
        return months

    def del_file_archive(self, file_name: str) -> None:
        '''Удаляет файл из архиа по его названию генерируя путь к архиву'''
        os.remove(os.path.join(self.path_archive_files, file_name))

    @staticmethod
    def get_number_by_chars(number: int | tuple) -> str:
        """Преобразует число в буквенное представление. Имеет два режима если передать целое число - просто вернет
        его буквенное представление, но при передаче картежа вернет представление с разделяющей точкой"""
        excel_col_name = lambda n: '' if n <= 0 else excel_col_name((n - 1) // 26) + chr((n - 1) % 26 + ord('A'))
        if isinstance(number, tuple):
            a, b = number[0], number[-1]
            return excel_col_name(a) + '.' + excel_col_name(b)

        return excel_col_name(number)

    @staticmethod
    def get_chars_by_number(chars: str) -> int | tuple:
        """Преобразует буквенное представление в число. Имеет два режима если передать буквенное представление без точки - просто вернет
                его числовое представление, но при передаче буквенного представление с разделяющей точкой вернет картеж из целых чисел"""
        excel_col_num = lambda a: 0 if a == '' else 1 + ord(a[-1]) - ord('A') + 26 * excel_col_num(a[:-1])
        if '.' in chars:
            a, b = chars.split('.')
            return excel_col_num(a), excel_col_num(b)
        return excel_col_num(chars)

    def save_f6(self, file_name='f6.xlsx'):
        BASE_COORDS = (0, 0)

        CELLS_INIT = {
            self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get(
                'specialization', ''),
            self.get_number_by_chars(21 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('group',
                                                                                                             ''),
            self.get_number_by_chars(24 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Кл.руководитель ' + (
                self.user.create_shorts_fio(self.user.parametrs.get('teamleader')) if self.user.parametrs.get(
                    'teamleader') else ''),
            self.get_number_by_chars(24 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): "Староста группы " + (
                self.user.create_shorts_fio(self.user.parametrs.get('offical_name')) if self.user.parametrs.get(
                    'offical_name') else ''),
            self.get_number_by_chars(35 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): 'Из них',
            self.get_number_by_chars(34 + BASE_COORDS[0]) + str(
                2 + BASE_COORDS[1]): f'за {self.MONTHS[self.period[0] - 1]} {str(self.period[1])}',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): '№',
            self.get_number_by_chars(36 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1]): 'по неув',
            self.get_number_by_chars(35 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1]): 'по ув',
            self.get_number_by_chars(34 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): 'Итого',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): 'ФИО',
            self.get_number_by_chars(18 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): 'Группа',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): 'Специальность',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(
                1 + BASE_COORDS[1]): 'ВЕДОМОСТЬ УЧЁТА ЧАСОВ, ПРОПУЩЕННЫХ СТУДЕНТАМИ',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Чел.Час',
            self.get_number_by_chars(6 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Посещ.об',
            self.get_number_by_chars(6 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Посящаемость Кач.',
            self.get_number_by_chars(17 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Прогул 1 студ.',
            self.get_number_by_chars(17 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): '=',
            self.get_number_by_chars(29 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1]): 'Итого',
        }
        CELLS_MERGE = [
            f'{self.get_number_by_chars(3 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(8 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(3 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(5 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(3 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(5 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(20 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(2 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(35 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(36 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(12 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(16 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(12 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(16 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(29 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])}:{self.get_number_by_chars(33 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}:{self.get_number_by_chars(36 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(34 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(36 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(15 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}:{self.get_number_by_chars(26 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(10 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(10 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(21 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(24 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}:{self.get_number_by_chars(33 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(24 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}:{self.get_number_by_chars(33 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])}',

        ]
        SET_WIDTH_CELLS = {
            self.get_number_by_chars(1 + BASE_COORDS[0]): 3,
            self.get_number_by_chars(2 + BASE_COORDS[0]): 30,
            self.get_number_by_chars(34 + BASE_COORDS[0]): 8,
            self.get_number_by_chars(35 + BASE_COORDS[0]): 8,
            self.get_number_by_chars(36 + BASE_COORDS[0]): 8,

        }

        statistics = self.get_statistics()
        if statistics:
            CELLS_INIT[self.get_number_by_chars(34 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get(
                'all_absence')
            CELLS_INIT[self.get_number_by_chars(35 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get(
                'sack_days_hours')
            CELLS_INIT[self.get_number_by_chars(36 + BASE_COORDS[0]) + str(35 + BASE_COORDS[1])] = statistics.get(
                'absence_days_hours')
            CELLS_INIT[self.get_number_by_chars(34 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])] = statistics.get(
                'all_hours')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get(
                'man_hours')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get(
                'total_attendance')
            CELLS_INIT[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get(
                'quality_attendance')
            CELLS_INIT[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get(
                'absences_by_student')

        wb = Workbook()
        ws = wb.active

        patternfull = PatternFill(fgColor='D9D9D9', fill_type='solid')
        font = Font(name='Calibri', size=16, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
        font_sick_days = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='single',
                              strike=False,
                              color='FF000000')
        font_absence_days = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                                 strike=False,
                                 color='FF000000')
        fill_border_style = Border(Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"),
                                   Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"))

        ws[self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])].border = fill_border_style
        ws[self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])].font = font
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[1]
        ws[self.get_number_by_chars(18 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[1]

        for name_call in (
                self.get_number_by_chars(34 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(35 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]),
                self.get_number_by_chars(34 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]),
                self.get_number_by_chars(35 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
        ):
            ws[name_call].border = Border(bottom=Side(style='thin'))

        # Настройка ширины столбцов и нумерация дней

        for i in range(3, 34):
            name_cell = self.get_number_by_chars(i + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])

            ws[name_cell] = i - 2
            ws.column_dimensions[self.get_number_by_chars(i + BASE_COORDS[0])].width = 3.5

            if i - 2 in self.days:
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])] = self.days.get(i - 2)

            else:
                ws[name_cell].fill = patternfull
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])] = '✖'

        for name, value in CELLS_INIT.items():
            ws[name] = value
            ws[name].alignment = Alignment(horizontal='center')

        for i in CELLS_MERGE:
            ws.merge_cells(i)

        for key, value in SET_WIDTH_CELLS.items():
            ws.column_dimensions[key].width = value

        students = self.students[:30]

        column = self.get_number_by_chars(2 + BASE_COORDS[0])

        for indx, student in enumerate(students):
            row = str(5 + indx + BASE_COORDS[1])
            fio, s_d, a_d = student.get_statistic_for_student().values()
            ws[column + row] = self.CLASS_STUDENT.create_shorts_fio(fio)
            ws[self.get_number_by_chars(self.get_chars_by_number(column) - 1) + row] = indx + 1
            for c in range(3, 34):
                if c - 2 in a_d[0]:
                    ws[self.get_number_by_chars(c + BASE_COORDS[0]) + row] = a_d[0].get(c - 2, '')
                    ws[self.get_number_by_chars(c + BASE_COORDS[0]) + row].font = font_absence_days

                elif c - 2 in s_d[0]:
                    ws[self.get_number_by_chars(c + BASE_COORDS[0]) + row] = s_d[0].get(c - 2, '')
                    ws[self.get_number_by_chars(c + BASE_COORDS[0]) + row].font = font_sick_days

            ws[self.get_number_by_chars(35 + BASE_COORDS[0]) + row] = s_d[1] if s_d[1] else ''
            ws[self.get_number_by_chars(36 + BASE_COORDS[0]) + row] = a_d[1] if a_d[1] else ''

        for i in range(1, 37):
            for j in range(3, 36):
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j + BASE_COORDS[1])].border = fill_border_style
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j + BASE_COORDS[1])].alignment = Alignment(
                    horizontal='center')
                if 3 <= i <= 33:
                    if i - 2 not in self.days:
                        ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j)].fill = patternfull

        wb.save(file_name)
        return file_name

    def save_f6_marks(self, file_name='f6marks.xlsx'):
        BASE_COORDS = (0, 0)

        CELLS_INIT = {
            self.get_number_by_chars(4 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get(
                'specialization', ''),
            self.get_number_by_chars(9 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): self.user.parametrs.get('group',
                                                                                                            ''),
            self.get_number_by_chars(10 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Кл.руководитель ' + (
                self.user.create_shorts_fio(self.user.parametrs.get('teamleader')) if self.user.parametrs.get(
                    'teamleader') else ''),
            self.get_number_by_chars(10 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): "Староста группы " + (
                self.user.create_shorts_fio(self.user.parametrs.get('offical_name')) if self.user.parametrs.get(
                    'offical_name') else ''),
            self.get_number_by_chars(12 + BASE_COORDS[0]) + str(
                2 + BASE_COORDS[1]): f'за {self.MONTHS[self.period[0] - 1]} {str(self.period[1])}',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): '№',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1]): 'ФИО',
            self.get_number_by_chars(8 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): 'Группа',
            self.get_number_by_chars(3 + BASE_COORDS[0]) + str(2 + BASE_COORDS[1]): 'Специальность',
            self.get_number_by_chars(1 + BASE_COORDS[0]) + str(
                1 + BASE_COORDS[1]): 'ВЕДОМОСТЬ УЧЁТА ЧАСОВ, УСПИВАЕМОСТИ СТУДЕНТАМИ',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Кол-во студ. им-х 2',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Кол-во студ. им-х 5',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1]): 'Кол-во студ. им-х одну 3',
            self.get_number_by_chars(2 + BASE_COORDS[0]) + str(39 + BASE_COORDS[1]): 'Кол-во студ. им-х 4 и 5',
            self.get_number_by_chars(5 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]): 'Успеваемость общая',
            self.get_number_by_chars(5 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]): 'Успеваемость качественная',

        }
        CELLS_MERGE = [
            f'{self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(2 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])}:{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])}',
            f'{self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}:{self.get_number_by_chars(13 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])}',

        ]
        SET_WIDTH_CELLS = {
            self.get_number_by_chars(1 + BASE_COORDS[0]): 3,
            self.get_number_by_chars(2 + BASE_COORDS[0]): 30,
        }
        statistics = self.get_statistics_marks()
        if statistics:
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])] = statistics.get(
                'heaving_2')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get(
                'heaving_5')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(38 + BASE_COORDS[1])] = statistics.get(
                'heaving_one_3')
            CELLS_INIT[self.get_number_by_chars(3 + BASE_COORDS[0]) + str(39 + BASE_COORDS[1])] = statistics.get(
                'heaving_4_and_5')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])] = statistics.get(
                'total_academic_performance')
            CELLS_INIT[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])] = statistics.get(
                'quality_academic_performance')

        wb = Workbook()
        ws = wb.active

        font = Font(name='Calibri', size=16, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
        fill_border_style = Border(Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"),
                                   Side(border_style="thin", color="000000"), Side(border_style="thin", color="000000"))

        ws[self.get_number_by_chars(2 + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])].border = fill_border_style
        ws[self.get_number_by_chars(1 + BASE_COORDS[0]) + str(1 + BASE_COORDS[1])].font = font
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]
        ws[self.get_number_by_chars(7 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1])].number_format = BUILTIN_FORMATS[10]

        for name_call in (
                self.get_number_by_chars(12 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]),
                self.get_number_by_chars(13 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(12 + BASE_COORDS[0]) + str(37 + BASE_COORDS[1]),
                self.get_number_by_chars(13 + BASE_COORDS[0]) + str(36 + BASE_COORDS[1]),
        ):
            ws[name_call].border = Border(bottom=Side(style='thin'))

        # Настройка ширины столбцов и нумерация дней

        for i in range(3, 14):
            name_call1 = self.get_number_by_chars(i + BASE_COORDS[0]) + str(3 + BASE_COORDS[1])
            name_cell2 = self.get_number_by_chars(i + BASE_COORDS[0]) + str(4 + BASE_COORDS[1])
            ws[name_call1] = self.couples.get(str(i - 2))[0] if self.couples.get(str(i - 2)) else ''
            ws[name_cell2] = self.CLASS_STUDENT.create_shorts_fio(self.couples.get(str(i - 2))[1]) if self.couples.get(
                str(i - 2)) else ''
            ws.column_dimensions[self.get_number_by_chars(i + BASE_COORDS[0])].width = 15

        for name, value in CELLS_INIT.items():
            ws[name] = value
            ws[name].alignment = Alignment(horizontal='center')

        for i in CELLS_MERGE:
            ws.merge_cells(i)

        for key, value in SET_WIDTH_CELLS.items():
            ws.column_dimensions[key].width = value

        students = self.students[:30]

        column = self.get_number_by_chars(2 + BASE_COORDS[0])

        for indx, student in enumerate(students):
            row = str(5 + indx + BASE_COORDS[1])
            fio, s_d, a_d = student.get_statistic_for_student().values()
            ws[column + row] = self.CLASS_STUDENT.create_shorts_fio(fio)
            ws[self.get_number_by_chars(self.get_chars_by_number(column) - 1) + row] = indx + 1

            for i in student.marks:
                ws[self.get_number_by_chars(i + 2) + row] = student.marks.get(i, '')

        for i in range(1, 14):
            for j in range(3, 35):
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j + BASE_COORDS[1])].border = fill_border_style
                ws[self.get_number_by_chars(i + BASE_COORDS[0]) + str(j + BASE_COORDS[1])].alignment = Alignment(
                    horizontal='center')
                ws.row_dimensions[j].height = 15

        wb.save(file_name)
        return file_name


if __name__ == '__main__':
    pass
