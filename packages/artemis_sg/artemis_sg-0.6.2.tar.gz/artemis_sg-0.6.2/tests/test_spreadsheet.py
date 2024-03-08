# ruff: noqa: S101
import logging
import os
import shutil
from unittest.mock import Mock

import pytest
from openpyxl import load_workbook

from artemis_sg import spreadsheet


@pytest.fixture()
def populated_target_directory(tmp_path_factory, image_filepath):
    path = tmp_path_factory.mktemp("data")
    shutil.copyfile(image_filepath, os.path.join(path, "9781680508604.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9780691025551.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9780691025552.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "672125069899.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9999999999990.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "9999999999990-1.jpg"))
    shutil.copyfile(image_filepath, os.path.join(path, "FI-1234.jpg"))
    with open(os.path.join(path, "9999999999999.jpg"), "w") as f:
        f.write("I am not an image file")
        f.close()
    yield path


def test_sheet_image_no_worksheet(
    caplog, spreadsheet_filepath, target_directory, populated_target_directory
):
    """
    GIVEN a spreadsheet
    AND an output filepath
    WHEN sheet_image is run in debug mode
    THEN a file is saved as the given output file.
    AND log shows images inserted
    """
    caplog.set_level(logging.INFO)
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = None
    outfile = os.path.join(target_directory, "sheet_image_output_file.xlsx")
    image = "9780691025551.jpg"
    filepath = os.path.join(populated_target_directory, image)

    spreadsheet.sheet_image(
        vendor_code, workbook, worksheet, populated_target_directory, outfile
    )

    assert os.path.exists(outfile)
    assert (
        "root",
        logging.INFO,
        f"spreadsheet.insert_image: Inserted '{filepath}'.",
    ) in caplog.record_tuples
    assert (
        "root",
        logging.INFO,
        "spreadsheet.sheet_image: Worksheet is Sheet1",
    ) in caplog.record_tuples


def test_sheet_image_output_file(
    caplog, spreadsheet_filepath, target_directory, populated_target_directory
):
    """
    GIVEN a spreadsheet
    AND the workbook and worksheet references for the spreadsheet
    AND an output filepath
    WHEN sheet_image is run in debug mode
    THEN a file is saved as the given output file.
    AND the data is in the expected order
    AND log shows images inserted
    """
    caplog.set_level(logging.INFO)
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"
    outfile = os.path.join(target_directory, "sheet_image_output_file.xlsx")
    image = "9780691025551.jpg"
    filepath = os.path.join(populated_target_directory, image)

    spreadsheet.sheet_image(
        vendor_code, workbook, worksheet, populated_target_directory, outfile
    )

    wb = load_workbook(outfile)
    ws = wb.worksheets[0]
    expected_sheet_keys = ["ISBN-13", "IMAGE", "ORDER", "FOO", "BAR", "BAZ"]
    sheet_keys = spreadsheet.get_sheet_keys(ws)

    assert os.path.exists(outfile)
    assert sheet_keys == expected_sheet_keys
    assert (
        "root",
        logging.INFO,
        f"spreadsheet.insert_image: Inserted '{filepath}'.",
    ) in caplog.record_tuples


def test_sheet_image_isbn_as_floats(
    caplog, spreadsheet_filepath, target_directory, populated_target_directory
):
    """
    GIVEN a spreadsheet containing ISBNs as floating point numbers
    AND the workbook and worksheet references for the spreadsheet
    AND an output filepath
    WHEN sheet_image is run in debug mode
    THEN a file is saved as the given output file.
    AND log ISBN inserted as integer
    """
    caplog.set_level(logging.INFO)
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"
    outfile = os.path.join(target_directory, "sheet_image_output_file.xlsx")
    image = "9780691025552.jpg"
    filepath = os.path.join(populated_target_directory, image)

    spreadsheet.sheet_image(
        vendor_code, workbook, worksheet, populated_target_directory, outfile
    )

    assert os.path.exists(outfile)
    assert (
        "root",
        logging.INFO,
        f"spreadsheet.insert_image: Inserted '{filepath}'.",
    ) in caplog.record_tuples


def test_mkthumbs_deletes_corrupted_image(populated_target_directory):
    """
    GIVEN a corrupted JPEG file in an image directory
    WHEN mkthumbs is run with the image directory
    THEN mkthumbs should complete without error
    AND the corrupted file should not exist in the image directory
    """
    corrupted_file = "9999999999999.jpg"
    image_directory = str(populated_target_directory)

    spreadsheet.mkthumbs(image_directory)

    assert True
    assert corrupted_file not in os.listdir(image_directory)


def test_mkthumbs_creates_thumbnails(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory
    WHEN mkthumbs is run with the image directory
    THEN thumbnails subdirectory should be created
    AND the JPEG should exist in the subdirectory
    """
    image_file = "9999999999990.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert subdir in os.listdir(image_directory)
    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_doesnt_create_supplementals(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory with a '-1' suffix
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should not exist in the subdirectory
    """
    image_file = "9999999999990-1.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file not in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_creates_invalid_isbn(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory named with an invalid isbn
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should exist in the subdirectory
    """
    image_file = "672125069899.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_creates_item_with_hyphen(populated_target_directory):
    """
    GIVEN a JPEG file in an image directory named with a hyphen ("FI-1234.jpg")
    WHEN mkthumbs is run with the image directory
    THEN the JPEG should exist in the subdirectory
    """
    image_file = "FI-1234.jpg"
    image_directory = str(populated_target_directory)
    subdir = "thumbnails"

    spreadsheet.mkthumbs(image_directory)

    assert image_file in os.listdir(os.path.join(image_directory, subdir))


def test_mkthumbs_ignores_no_basename(tmp_path_factory, image_filepath):
    """
    GIVEN a JPEG file in an image directory named '.jpg'
    WHEN mkthumbs is run with the image directory
    THEN the file should not exist in the subdirectory
    """
    subdir = "thumbnails"
    image_file = ".jpg"
    image_directory = tmp_path_factory.mktemp("test_no_basename")
    shutil.copyfile(image_filepath, os.path.join(image_directory, image_file))

    spreadsheet.mkthumbs(image_directory)

    assert image_file not in os.listdir(os.path.join(image_directory, subdir))


def test_get_order_items(spreadsheet_filepath):
    """
    GIVEN a spreadsheet with "ISBN-13" and "Order" columns
    AND the spreadsheet contains rows with items and quantities
    WHEN get_order_items is run with a vendor_code
    AND the workbook and worksheet references for the spreadsheet
    THEN a list of order items is returned
    """
    expected_list = [
        ("9780300153750", "42"),
        ("9780691025551", "3"),
        ("9780691025552", "3"),
    ]
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"

    order_items = spreadsheet.get_order_items(vendor_code, workbook, worksheet)

    assert isinstance(order_items, list)
    assert order_items == expected_list


def test_get_sheet_data_gdoc_id_no_worksheet(monkeypatch):
    """
    Given a valid Google Doc ID
    WHEN get_sheet_data is executed with the ID
    AND no worksheet name is provided
    THEN the first worksheet is chosen
    """
    workbook = "not_a_file"
    worksheet = None
    expected_worksheet = "CoolSheet"

    mock_sheets = Mock()
    call_chain = (
        "spreadsheets.return_value."
        "get.return_value."
        "execute.return_value."
        "get.return_value."
        "pop.return_value.get.return_value.get.return_value"
    )
    config = {"name": "mock_sheets", call_chain: expected_worksheet}
    mock_sheets.configure_mock(**config)
    monkeypatch.setattr(
        spreadsheet.app_creds, "app_creds", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(spreadsheet, "build", lambda *args, **kwargs: mock_sheets)

    spreadsheet.get_sheet_data(workbook, worksheet)

    mock_sheets.spreadsheets().get.assert_called_with(spreadsheetId=workbook)
    mock_sheets.spreadsheets().values().get.assert_called_with(
        range=expected_worksheet, spreadsheetId=workbook
    )


def test_get_sheet_data_file_worksheet(monkeypatch, spreadsheet_filepath):
    """
    Given a valid spreadsheet file
    WHEN get_sheet_data is executed with the file path
    AND no worksheet name is provided
    THEN the first worksheet data is loaded
    """
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"

    data = spreadsheet.get_sheet_data(workbook, worksheet)

    assert isinstance(data, list)
    assert "ISBN-13" in data[0]


def test_get_sheet_data_file_no_worksheet(monkeypatch, spreadsheet_filepath):
    """
    Given a valid spreadsheet file
    WHEN get_sheet_data is executed with the file path
    AND no worksheet name is provided
    THEN the first worksheet data is loaded
    """
    workbook = spreadsheet_filepath
    worksheet = None

    data = spreadsheet.get_sheet_data(workbook, worksheet)

    assert isinstance(data, list)
    assert "ISBN-13" in data[0]


def test_sheet_waves_output_file(
    spreadsheet_filepath, target_directory, valid_datafile
):
    """
    GIVEN a spreadsheet
    AND the workbook and worksheet references for the spreadsheet
    AND an output filepath
    WHEN sheet_waves is run in debug mode
    THEN a file is saved as the given output file.
    AND the file contains a Description column
    AND the file contains a Dimension column
    AND the file contains a ImageURL0 column
    AND the file contains a ImageURL1 column
    AND the file contains a ImageURL2 column
    AND the file contains a ImageURL3 column
    AND the file contains a ImageURL4 column
    AND the file contains a ImageURL5 column
    AND the file contains a ImageURL6 column
    """
    vendor_code = "sample"
    workbook = spreadsheet_filepath
    worksheet = "Sheet1"
    outfile = os.path.join(target_directory, "sheet_waves_output_file.xlsx")
    spreadsheet.sheet_waves(
        vendor_code, workbook, worksheet, outfile, valid_datafile
    )

    assert os.path.exists(outfile)
    wb = spreadsheet.load_workbook(outfile)
    ws = spreadsheet.get_worksheet(wb, worksheet)
    row01 = ws[1]
    headers = [ cell.value for cell in row01 ]
    assert "Description" in headers
    assert "Dimension" in headers
    assert "ImageURL0" in headers
    assert "ImageURL1" in headers
    assert "ImageURL2" in headers
    assert "ImageURL3" in headers
    assert "ImageURL4" in headers
    assert "ImageURL5" in headers
    assert "ImageURL6" in headers
