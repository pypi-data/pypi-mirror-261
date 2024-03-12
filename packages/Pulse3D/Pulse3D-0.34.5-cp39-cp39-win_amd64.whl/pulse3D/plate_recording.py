# -*- coding: utf-8 -*-
from collections import defaultdict
import datetime
import glob
import json
import os
import tempfile
from typing import Any
from typing import Dict
from typing import List
from typing import Optional
from typing import Union
import uuid
import zipfile

import h5py
from mantarray_magnet_finding.exceptions import UnableToConvergeError
from mantarray_magnet_finding.utils import calculate_magnetic_flux_density_from_memsic
from nptyping import NDArray
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from scipy import interpolate
from semver import VersionInfo
import structlog
from xlsxwriter.utility import xl_cell_to_rowcol

from .compression_cy import compress_filtered_magnetic_data
from .constants import *
from .exceptions import DuplicateWellsFoundError
from .exceptions import IncorrectOpticalFileFormatError
from .exceptions import NoRecordingFilesLoadedError
from .exceptions import SubprotocolFormatIncompatibleWithInterpolationError
from .magnet_finding import find_magnet_positions
from .magnet_finding import fix_dropped_samples
from .magnet_finding import format_well_file_data
from .stimulation import aggregate_timepoints
from .stimulation import create_stim_session_waveforms
from .stimulation import realign_interpolated_stim_data
from .transforms import apply_empty_plate_calibration
from .transforms import apply_noise_filtering
from .transforms import apply_sensitivity_calibration
from .transforms import calculate_displacement_from_voltage
from .transforms import calculate_force_from_displacement
from .transforms import calculate_voltage_from_gmr
from .transforms import create_filter
from .transforms import noise_cancellation
from .utils import get_experiment_id
from .utils import get_stiffness_factor
from .utils import get_well_name_from_h5
from .utils import truncate

log = structlog.getLogger()


class MantarrayH5FileCreator(h5py.File):
    """Creates an H5 file with the basic format/layout."""

    def __init__(
        self, file_name: str, file_format_version: str = CURRENT_BETA2_HDF5_FILE_FORMAT_VERSION
    ) -> None:
        super().__init__(
            file_name,
            "w",
            # tried to specify this ('earliest', 'v110') to be more backward compatible
            # but it didn't work for unknown reasons (gave error when trying to set swmr_mode=True)
            libver="latest",
            userblock_size=512,  # minimum size is 512 bytes
        )

        self.attrs[FILE_FORMAT_VERSION_METADATA_KEY] = file_format_version


class WellFile:
    def __init__(
        self,
        file_path: str,
        sampling_period: Optional[Union[int, float]] = None,
        # TODO unit test the stiffness factor (auto and override)
        stiffness_factor: Optional[int] = None,
        has_inverted_post_magnet: bool = False,
    ):
        self.displacement: NDArray[(2, Any), np.float64]
        self.force: NDArray[(2, Any), np.float64]
        self.stim_sessions: List[NDArray[(2, Any), int]] = []

        if stiffness_factor not in (*POST_STIFFNESS_OVERRIDE_OPTIONS, None):
            raise ValueError(
                f"Invalid Post Stiffness {stiffness_factor}, must be in {POST_STIFFNESS_OVERRIDE_OPTIONS}"
            )

        self.has_inverted_post_magnet = has_inverted_post_magnet
        self.stiffness_override = stiffness_factor is not None

        if file_path.endswith(".h5"):
            self.is_magnetic_data = True

            self._load_data_from_h5_file(file_path)

            self.tissue_sampling_period = (
                sampling_period if sampling_period else self[TISSUE_SAMPLING_PERIOD_UUID]
            )

            if self.stiffness_override:
                self.stiffness_factor = stiffness_factor
            elif not self.get(IS_CALIBRATION_FILE_UUID, False):
                # earlier versions of files do not have the IS_CALIBRATION_FILE_UUID in their metadata
                experiment_id = get_experiment_id(self[PLATE_BARCODE_UUID])
                self.stiffness_factor = get_stiffness_factor(experiment_id, self[WELL_NAME_UUID])
            else:
                # calibration recordings do not have an associated barcode or post stiffness since they
                # are creatd when a plate is not even on the istrument, so just set the stiffness factor to 1
                self.stiffness_factor = CALIBRATION_STIFFNESS_FACTOR

            if self.version < VersionInfo.parse("1.0.0"):  # type: ignore
                # Ref data not yet added to Beta 2 files
                self[UTC_FIRST_REF_DATA_POINT_UUID] = self._extract_datetime(UTC_FIRST_REF_DATA_POINT_UUID)
                # setup noise filter, should probably add beta 2 file support for this
                self.noise_filter_uuid = (
                    TSP_TO_DEFAULT_FILTER_UUID[self.tissue_sampling_period] if self.is_magnetic_data else None
                )
                self.filter_coefficients = (
                    create_filter(self.noise_filter_uuid, self.tissue_sampling_period)
                    if self.noise_filter_uuid
                    else None
                )

                self._load_magnetic_data()
            else:
                if self.has_inverted_post_magnet:
                    raise ValueError("V1 algorithm will automatically fix this issue")

                self.noise_filter_uuid = None
                self.filter_coefficients = None
                self.is_magnetic_data = False

            # extract datetime metadata
            for uuid_ in (
                UTC_BEGINNING_RECORDING_UUID,
                UTC_BEGINNING_DATA_ACQUISTION_UUID,
                UTC_FIRST_TISSUE_DATA_POINT_UUID,
            ):
                self[uuid_] = self._extract_datetime(uuid_)

            # format platemap metadata
            for uuid_ in (PLATEMAP_NAME_UUID, PLATEMAP_LABEL_UUID):
                val = self.get(uuid_, NOT_APPLICABLE_LABEL)
                if val == str(NOT_APPLICABLE_H5_METADATA):
                    val = NOT_APPLICABLE_LABEL
                self[uuid_] = val

        elif file_path.endswith(".xlsx"):
            self._excel_sheet = _get_single_sheet(file_path)
            self.file_name = os.path.basename(file_path)
            self.attrs = {k: v for k, v in _load_optical_file_attrs(self._excel_sheet).items()}
            self.version = self[FILE_FORMAT_VERSION_METADATA_KEY]
            self.is_magnetic_data = False
            self.stiffness_factor = None

            for uuid_ in (PLATEMAP_NAME_UUID, PLATEMAP_LABEL_UUID):
                self[uuid_] = NOT_APPLICABLE_LABEL
            # skip all other transforms
            self.force = self[TISSUE_SENSOR_READINGS].copy()
            # timepoints still need to be in µs
            self.force[0] *= MICRO_TO_BASE_CONVERSION

    def _load_data_from_h5_file(self, file_path: str) -> None:
        with h5py.File(file_path, "r") as h5_file:
            self.file_name = os.path.basename(h5_file.filename)
            self.attrs = {attr: h5_file.attrs[attr] for attr in list(h5_file.attrs)}
            self.version = self[FILE_FORMAT_VERSION_METADATA_KEY]

            if self.version < VersionInfo.parse("1.0.0"):
                # load sensor data. This is only possible to do here for Beta 1 data files
                self[TISSUE_SENSOR_READINGS] = self._load_reading(h5_file, TISSUE_SENSOR_READINGS)
                self[REFERENCE_SENSOR_READINGS] = self._load_reading(h5_file, REFERENCE_SENSOR_READINGS)
            else:
                for dataset in (
                    TIME_INDICES,
                    TIME_OFFSETS,
                    TISSUE_SENSOR_READINGS,
                    REFERENCE_SENSOR_READINGS,
                    STIMULATION_READINGS,
                ):
                    self[dataset] = h5_file[dataset][:]

    def _load_reading(self, h5_file, reading_type: str) -> NDArray[(Any, Any), int]:
        sampling_period = self[
            REF_SAMPLING_PERIOD_UUID
            if reading_type == REFERENCE_SENSOR_READINGS
            else TISSUE_SAMPLING_PERIOD_UUID
        ]
        time_step = int(sampling_period / MICROSECONDS_PER_CENTIMILLISECOND)

        tissue_contraction_amplitudes = h5_file[reading_type][:]
        num_data_points = tissue_contraction_amplitudes.shape[-1]

        timepoints = np.arange(num_data_points) * time_step

        return np.array([timepoints, tissue_contraction_amplitudes], dtype=np.int32)

    def _load_magnetic_data(self):
        adj_raw_tissue_reading = self[TISSUE_SENSOR_READINGS].copy()

        time_conversion = (
            MICROSECONDS_PER_CENTIMILLISECOND if self.is_magnetic_data else MICRO_TO_BASE_CONVERSION
        )
        adj_raw_tissue_reading[0] *= time_conversion

        # magnetic data is flipped
        if self.is_magnetic_data:
            adj_raw_tissue_reading[1] *= -1

        # some posts have a flipped magnet
        if self.has_inverted_post_magnet:
            adj_raw_tissue_reading[1] *= -1

        self.raw_tissue_magnetic_data: NDArray[(2, Any), int] = adj_raw_tissue_reading
        self.raw_reference_magnetic_data: NDArray[(2, Any), int] = self[REFERENCE_SENSOR_READINGS].copy()

        self.sensitivity_calibrated_tissue_gmr: NDArray[(2, Any), int] = apply_sensitivity_calibration(
            self.raw_tissue_magnetic_data
        )

        self.sensitivity_calibrated_reference_gmr: NDArray[(2, Any), int] = apply_sensitivity_calibration(
            self.raw_reference_magnetic_data
        )

        self.noise_cancelled_magnetic_data: NDArray[(2, Any), int] = noise_cancellation(
            self.sensitivity_calibrated_tissue_gmr, self.sensitivity_calibrated_reference_gmr
        )

        self.fully_calibrated_magnetic_data: NDArray[(2, Any), int] = apply_empty_plate_calibration(
            self.noise_cancelled_magnetic_data
        )

        self.noise_filtered_magnetic_data: NDArray[(2, Any), int]
        if self.noise_filter_uuid is None:
            self.noise_filtered_magnetic_data = self.fully_calibrated_magnetic_data
        else:
            self.noise_filtered_magnetic_data = apply_noise_filtering(
                self.fully_calibrated_magnetic_data, self.filter_coefficients
            )

        self.compressed_magnetic_data: NDArray[(2, Any), int] = compress_filtered_magnetic_data(
            self.noise_filtered_magnetic_data
        )
        self.compressed_voltage: NDArray[(2, Any), np.float32] = calculate_voltage_from_gmr(
            self.compressed_magnetic_data
        )
        self.compressed_displacement: NDArray[(2, Any), np.float32] = calculate_displacement_from_voltage(
            self.compressed_voltage
        )
        self.compressed_force: NDArray[(2, Any), np.float32] = calculate_force_from_displacement(
            self.compressed_displacement, stiffness_factor=self.stiffness_factor, in_mm=False
        )

        self.voltage: NDArray[(2, Any), np.float32] = calculate_voltage_from_gmr(
            self.noise_filtered_magnetic_data
        )
        self.displacement: NDArray[(2, Any), np.float64] = calculate_displacement_from_voltage(self.voltage)
        self.force: NDArray[(2, Any), np.float64] = calculate_force_from_displacement(
            self.displacement, stiffness_factor=self.stiffness_factor, in_mm=False
        )

    def get(self, key, default=None) -> Any:
        try:
            return self[key]
        except Exception:
            return default

    def __contains__(self, key):
        key = str(key) if isinstance(key, uuid.UUID) else key
        return key in self.attrs

    def __setitem__(self, key, newvalue):
        key = str(key) if isinstance(key, uuid.UUID) else key
        self.attrs[key] = newvalue

    def __getitem__(self, i):
        i = str(i) if isinstance(i, uuid.UUID) else i
        return self.attrs[i]

    def _extract_datetime(self, metadata_uuid: uuid.UUID) -> datetime.datetime:
        if self.version.split(".") < VersionInfo.parse("0.2.1"):
            if metadata_uuid == UTC_BEGINNING_RECORDING_UUID:
                """The use of this proxy value is justified by the fact that
                there is a 15 second delay between when data is recorded and
                when the GUI displays it, and because the GUI will send the
                timestamp of when the recording button is pressed."""
                acquisition_timestamp_str = self[UTC_BEGINNING_DATA_ACQUISTION_UUID]

                begin_recording = datetime.datetime.strptime(
                    acquisition_timestamp_str, DATETIME_STR_FORMAT
                ).replace(tzinfo=datetime.timezone.utc) + datetime.timedelta(seconds=15)

                return begin_recording
            if metadata_uuid == UTC_FIRST_TISSUE_DATA_POINT_UUID:
                """Early file versions did not include this metadata under a
                UUID, so we have to use this string identifier instead."""
                metadata_name = "UTC Timestamp of Beginning of Recorded Tissue Sensor Data"
                timestamp_str = self[metadata_name]

                return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
                    tzinfo=datetime.timezone.utc
                )
            if metadata_uuid == UTC_FIRST_REF_DATA_POINT_UUID:
                """Early file versions did not include this metadata under a
                UUID, so we have to use this string identifier instead."""
                timestamp_str = self["UTC Timestamp of Beginning of Recorded Reference Sensor Data"]

                return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
                    tzinfo=datetime.timezone.utc
                )

        timestamp_str = self[metadata_uuid]
        return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
            tzinfo=datetime.timezone.utc
        )


class PlateRecording:
    def __init__(
        self,
        path,
        recording_df: pd.DataFrame = None,
        start_time: Union[float, int] = 0,
        end_time: Optional[Union[float, int]] = None,
        # TODO unit test the stiffness factor (auto and override), inverted_post_magnet_wells
        stiffness_factor: Optional[int] = None,
        inverted_post_magnet_wells: Optional[List[str]] = None,
        well_groups: Optional[Dict[str, List[str]]] = None,
    ):
        self.path = path
        self.wells: List[WellFile] = []
        self._iter = 0
        # these may get overwritten later
        self.is_optical_recording = False
        self.contains_stim_data = False

        # Tanner (11/16/22): due to the needs of the scientists for the full analysis,
        # these params should only be used in the recording snapshot.
        # These params also have no effect on Beta 1 data.
        if start_time < 0:
            raise ValueError("'start_time' must be >= 0")
        if end_time and start_time >= end_time:
            raise ValueError("'start_time' must be < 'end_time'")
        self.start_time_secs = start_time
        self.end_time_secs = end_time

        self._created_from_dataframe = recording_df is not None

        if self.path.endswith(".zip"):
            with tempfile.TemporaryDirectory() as tmpdir:
                zf = zipfile.ZipFile(path)
                zf.extractall(path=tmpdir)

                if glob.glob(os.path.join(tmpdir, "**", "*.h5"), recursive=True):
                    self.wells, calibration_recordings = load_files(
                        tmpdir, stiffness_factor, inverted_post_magnet_wells
                    )
                elif xlsx_files := glob.glob(os.path.join(tmpdir, "**", "*.xlsx"), recursive=True):
                    self._load_optical_well_files(xlsx_files, stiffness_factor)
        elif self.path.endswith(".xlsx"):  # optical file
            self._load_optical_well_files([self.path], stiffness_factor)
        else:  # .h5 files
            self.wells, calibration_recordings = load_files(
                self.path, stiffness_factor, inverted_post_magnet_wells
            )

        # make sure at least one WellFile was loaded
        if not any(self.wells):
            raise NoRecordingFilesLoadedError()

        if len(self.wells) > len(set(w[WELL_NAME_UUID] for w in self.wells)):
            raise DuplicateWellsFoundError()

        # ensure wells are in correct order A1, B1,.., A2, B2,...
        self.wells.sort(key=lambda w: (int(w[WELL_NAME_UUID][1:]), w[WELL_NAME_UUID][0]))

        # set up platemap info
        first_avaliable_well = next(iter(self))
        self.platemap_name = first_avaliable_well[PLATEMAP_NAME_UUID]
        platemap_labels = defaultdict(list)

        for well_file in self:
            if well_groups is None:
                label = well_file[PLATEMAP_LABEL_UUID]
                # only add to platemap_labels if label has been assigned
                if label != NOT_APPLICABLE_LABEL:
                    platemap_labels[label].append(well_file[WELL_NAME_UUID])
            else:
                # default all labels to NA first
                well_file[PLATEMAP_LABEL_UUID] = NOT_APPLICABLE_LABEL
                for label, well_names in well_groups.items():
                    if well_file[WELL_NAME_UUID] in well_names:
                        well_file[PLATEMAP_LABEL_UUID] = label
                        platemap_labels[label].append(well_file[WELL_NAME_UUID])

        self.platemap_labels = dict(platemap_labels)

        # currently file versions 1.0.0 and above must have all their data processed together
        if not self.is_optical_recording and self.wells[0].version >= VersionInfo.parse("1.0.0"):
            if self._created_from_dataframe:
                self._load_dataframe(recording_df)
            else:
                self._process_plate_data(calibration_recordings)

                if self.wells[0][FILE_FORMAT_VERSION_METADATA_KEY] >= VersionInfo.parse(
                    MIN_FILE_VERSION_FOR_STIM_INTERPOLATION
                ):
                    self._process_stim_data()

                self._handle_removal_of_initial_padding()

            self.contains_stim_data = any(wf.stim_sessions for wf in self)

    def _process_plate_data(self, calibration_recordings):
        if not all(isinstance(well_file, WellFile) for well_file in self.wells) or len(self.wells) != 24:
            raise NotImplementedError("All 24 wells must have a recording file present")

        if (
            not all(isinstance(well_file, WellFile) for well_file in calibration_recordings)
            or len(calibration_recordings) != 24
        ):
            raise NotImplementedError("All 24 wells must have a calibration file present")

        initial_magnet_finding_params = json.loads(
            self.wells[0].get(INITIAL_MAGNET_FINDING_PARAMS_UUID, r"{}")
        )

        sampling_period_us = self.wells[0][TISSUE_SAMPLING_PERIOD_UUID]
        sampling_freq = MICRO_TO_BASE_CONVERSION / sampling_period_us
        start_idx = int(self.start_time_secs * sampling_freq)
        end_idx = int(self.end_time_secs * sampling_freq) if self.end_time_secs else None
        analysis_window = slice(start_idx, end_idx)

        # load tissue data
        plate_data_array = format_well_file_data(self.wells)[:, analysis_window]
        fixed_plate_data_array = fix_dropped_samples(plate_data_array)
        plate_data_array_mt = calculate_magnetic_flux_density_from_memsic(fixed_plate_data_array)
        # load 'calibration' data
        baseline_data = format_well_file_data(calibration_recordings)
        baseline_data_mt = calculate_magnetic_flux_density_from_memsic(baseline_data)

        # create baseline data array
        baseline_data_mt = np.mean(baseline_data_mt[:, -BASELINE_MEAN_NUM_DATA_POINTS:], axis=1)

        try:
            # pass data into magnet finding alg
            log.info("Estimating magnet positions")
            estimated_magnet_positions = find_magnet_positions(
                plate_data_array_mt, baseline_data_mt, initial_magnet_finding_params
            )
        except UnableToConvergeError:
            log.exception("Unable to converge due to bad quality of data")
            raise

        flip_data = self.wells[0].version >= VersionInfo.parse("1.1.0")

        # create displacement and force arrays for each WellFile
        log.info("Create diplacement and force data for each well")
        for well_idx, well_file in enumerate(self):
            x = estimated_magnet_positions["X"][:, well_idx]
            if flip_data:
                x *= -1

            # have time indices start at 0
            time_indices = well_file[TIME_INDICES]
            adjusted_time_indices = time_indices[analysis_window] - time_indices[start_idx]

            well_file.displacement = np.array([adjusted_time_indices, x])

            well_file.force = calculate_force_from_displacement(
                well_file.displacement, stiffness_factor=well_file.stiffness_factor
            )

    def _process_stim_data(self) -> None:
        log.info("Interpolating stim sessions")

        start_time_us = int(self.start_time_secs * MICRO_TO_BASE_CONVERSION)
        end_time_us = (
            int(self.end_time_secs * MICRO_TO_BASE_CONVERSION)
            if self.end_time_secs
            else self.wells[0][TIME_INDICES][-1]
        )

        for wf in self:
            if not wf[STIMULATION_READINGS].shape[-1]:
                continue

            stim_protocol = json.loads(wf[STIMULATION_PROTOCOL_UUID])
            try:
                stim_sessions_waveforms = create_stim_session_waveforms(
                    stim_protocol["subprotocols"], wf[STIMULATION_READINGS], start_time_us, end_time_us
                )

            except SubprotocolFormatIncompatibleWithInterpolationError:
                log.exception("Subprotocol format not supported by interpolation")
                return

            is_voltage = stim_protocol["stimulation_type"] == "V"
            charge_conversion_factor = 1 if is_voltage else MILLI_TO_BASE_CONVERSION

            for waveform in stim_sessions_waveforms:
                if not waveform.shape[-1]:
                    continue
                waveform[0] -= wf[TIME_INDICES][0]
                waveform[1] /= charge_conversion_factor
                wf.stim_sessions.append(waveform)

    def _handle_removal_of_initial_padding(self) -> None:
        num_us_to_trim_from_start = self.wells[0].get(NUM_INITIAL_MICROSECONDS_TO_REMOVE_UUID)
        if not num_us_to_trim_from_start:
            return

        for well_file in self:
            well_file.displacement = well_file.displacement[
                :, well_file.displacement[0] >= num_us_to_trim_from_start
            ]
            well_file.force = well_file.force[:, well_file.force[0] >= num_us_to_trim_from_start]

            shift_amount = well_file.force[0, 0]

            well_file.force -= shift_amount
            well_file.displacement -= shift_amount
            for stim_session_arr in well_file.stim_sessions:
                stim_session_arr[0] -= shift_amount

    def _load_dataframe(self, df: pd.DataFrame) -> None:
        """Add time and force data to well files in PlateRecording.

        Args:
            df: pd.Dataframe existing time force data to be added
        """
        force_timepoints = df["Time (s)"].values

        try:
            stim_timepoints = df["Stim Time (µs)"].values
        except KeyError:
            stim_timepoints = None

        for wf in self.wells:
            well_name = wf[WELL_NAME_UUID]
            raw_force_amplitudes = df[f"{well_name}__raw"]
            wf.force = np.vstack((force_timepoints, raw_force_amplitudes)).astype(np.float64)
            wf.force = wf.force[:, ~np.isnan(raw_force_amplitudes)]

            if stim_timepoints is None:
                continue

            stim_col_titles = sorted([col for col in df if f"{well_name}__stim" in col])

            for col_title in stim_col_titles:
                stim_session_raw = np.array([stim_timepoints, df[col_title]])
                stim_session = stim_session_raw[:, ~np.isnan(stim_session_raw[1])].astype(int)
                wf.stim_sessions.append(stim_session)

    def _load_optical_well_files(self, file_paths: List[str], stiffness_factor: Union[int, None]):
        self.is_optical_recording = True

        for xlsx_path in file_paths:
            # check if xlsx is correct format and not pulse3d output file
            if _get_num_of_sheets(xlsx_path) > 1:
                raise IncorrectOpticalFileFormatError(
                    f"Incorrect number of sheets found for file {os.path.basename(xlsx_path)}"
                )

            well_file = WellFile(xlsx_path, stiffness_factor=stiffness_factor)
            self.wells.append(well_file)

    def to_dataframe(self, include_stim_data=True) -> pd.DataFrame:
        """Creates DataFrame from PlateRecording with all the data
        interpolated, normalized, and scaled.

        The returned dataframe contains one column for time in ms and
        one column for each well.
        """
        # get first valid well and set interpolation period. Creating new iter to be safe
        first_well = next(iter(self))

        # add interpolated force timepoints
        if self._created_from_dataframe:
            raise NotImplementedError("Cannot export a DF if created from a DF. Just use the original")

        min_time = min([wf.force[0, 0] for wf in self])
        max_time = max([wf.force[0, -1] for wf in self])

        interp_period = (
            first_well[INTERPOLATION_VALUE_UUID] if self.is_optical_recording else INTERPOLATED_DATA_PERIOD_US
        )
        interp_timepoints = np.arange(min_time, max_time + interp_period, interp_period)

        data = {"Time (s)": pd.Series(interp_timepoints)}

        # only attempt to output stim data if the file supports it and the caller requests it
        attempt_to_output_stim_data = (
            not self.is_optical_recording
            and first_well.version >= VersionInfo.parse(MIN_FILE_VERSION_FOR_STIM_INTERPOLATION)
            and include_stim_data
        )

        # add stim timepoints
        aggregate_stim_timepoints_us = None
        if attempt_to_output_stim_data:
            aggregate_stim_timepoints_us = aggregate_timepoints(
                [session_data[0] for wf in self for session_data in wf.stim_sessions]
            )
            aggregate_stim_timepoints_us_for_plotting = np.repeat(aggregate_stim_timepoints_us, 2)

        # only outputting stim data if an attempt to output stim data was made and the file actually has stim data in it
        is_outputting_stim_data = (
            aggregate_stim_timepoints_us is not None and aggregate_stim_timepoints_us.any()
        )
        if is_outputting_stim_data:
            data["Stim Time (µs)"] = pd.Series(aggregate_stim_timepoints_us_for_plotting)

        # iterating over self.wells instead of using __iter__ so well_idx is preserved
        for wf in self.wells:
            if not wf:
                continue

            well_name = wf.get(WELL_NAME_UUID)

            # add raw force data
            data[f"{well_name}__raw"] = pd.Series(wf.force[1, :])

            # add unit adjusted + normalized force data
            start_idx, end_idx = truncate(
                source_series=interp_timepoints, lower_bound=wf.force[0, 0], upper_bound=wf.force[0, -1]
            )

            interp_fn = interpolate.interp1d(wf.force[0, :], wf.force[1, :])
            interp_force_unewtons = interp_fn(interp_timepoints[start_idx : end_idx + 1])

            min_value = min(interp_force_unewtons)

            interp_force_newtons_normalized = (interp_force_unewtons - min_value) * MICRO_TO_BASE_CONVERSION
            data[well_name] = pd.Series(interp_force_newtons_normalized)

            # add stim data
            if is_outputting_stim_data:
                for i, session_data in enumerate(wf.stim_sessions):
                    data[f"{well_name}__stim_{i}"] = pd.Series(
                        realign_interpolated_stim_data(
                            aggregate_stim_timepoints_us_for_plotting, session_data
                        )
                    )

        df = pd.DataFrame(data)
        if not is_outputting_stim_data:
            df.dropna(inplace=True)

        return df

    @staticmethod
    def from_dataframe(path, **kwargs):
        # only allowed for one recording at a time assuming a user would only ever pass a dataframe to one recording
        log.info(f"Loading recording from file {os.path.basename(path)}")
        yield PlateRecording(path, **kwargs)

    @staticmethod
    def from_directory(path, **kwargs):
        # multi zip files
        for zf in glob.glob(os.path.join(path, "*.zip"), recursive=True):
            log.info(f"Loading recording from file {zf}")
            yield PlateRecording(zf, **kwargs)

        # multi optical files
        for of in glob.glob(os.path.join(path, "*.xlsx"), recursive=True):
            log.info(f"Loading optical data from file {of}")
            yield PlateRecording(of, **kwargs)

        # directory of .h5 files
        for dir in glob.glob(os.path.join(path, "*"), recursive=True):
            if glob.glob(os.path.join(dir, "*.h5"), recursive=True):
                yield PlateRecording(dir, **kwargs)

    def __iter__(self):
        self._iter = 0
        return self

    def __next__(self):
        while self._iter < len(self.wells):
            value = self.wells[self._iter]
            self._iter += 1
            if not value:
                continue
            return value
        else:
            raise StopIteration


# helpers
def load_files(
    path: str, stiffness_factor: Optional[int], inverted_post_magnet_wells: Optional[List[str]] = None
):
    if not inverted_post_magnet_wells:
        inverted_post_magnet_wells = []

    h5_files = glob.glob(os.path.join(path, "**", "*.h5"), recursive=True)

    recording_files = [f for f in h5_files if "Calibration" not in f]
    calibration_files = [f for f in h5_files if "Calibration" in f]

    tissue_well_files = [None] * len(recording_files)
    baseline_well_files = [None] * len(calibration_files)

    for f in recording_files:
        log.info(f"Loading data from {os.path.basename(f)}")
        well_name = get_well_name_from_h5(f)
        well_file = WellFile(
            f,
            stiffness_factor=stiffness_factor,
            has_inverted_post_magnet=well_name in inverted_post_magnet_wells,
        )
        tissue_well_files[well_file[WELL_INDEX_UUID]] = well_file  # type: ignore

    for f in calibration_files:
        log.info(f"Loading calibration data from {os.path.basename(f)}")
        well_file = WellFile(f, stiffness_factor=stiffness_factor)
        baseline_well_files[well_file[WELL_INDEX_UUID]] = well_file  # type: ignore

    return tissue_well_files, baseline_well_files


def _get_col_as_array(sheet: Worksheet, zero_based_row: int, zero_based_col: int) -> NDArray[(2, Any), float]:
    col_array = []
    result = _get_cell_value(sheet, zero_based_row, zero_based_col)
    zero_based_row += 1
    while result:
        col_array.append(float(result))
        result = _get_cell_value(sheet, zero_based_row, zero_based_col)
        zero_based_row += 1
    return np.array(col_array)


def _get_single_sheet(file_name: str) -> Any:
    work_book = load_workbook(file_name)
    return work_book[work_book.sheetnames[0]]


def _get_num_of_sheets(file_path: str) -> Any:
    work_book = load_workbook(file_path)
    return len(work_book.sheetnames)


def _get_cell_value(sheet: Worksheet, zero_based_row: int, zero_based_col: int) -> Optional[str]:
    result = sheet.cell(row=zero_based_row + 1, column=zero_based_col + 1).value
    if result is None:
        return result
    return str(result)


def _get_excel_metadata_value(sheet: Worksheet, metadata_uuid: uuid.UUID) -> Optional[str]:
    """Return a user-entered metadata value."""
    metadata_description = METADATA_UUID_DESCRIPTIONS[metadata_uuid]
    cell_name = EXCEL_OPTICAL_METADATA_CELLS.get(metadata_uuid, None)
    if cell_name is None:
        raise NotImplementedError(
            f"Metadata value for {metadata_description} is not contained in excel files of well data"
        )
    row, col = xl_cell_to_rowcol(cell_name)
    result = _get_cell_value(sheet, row, col)
    return result


def _load_optical_file_attrs(sheet: Worksheet):
    raw_tissue_reading = np.array((_get_col_as_array(sheet, 1, 0), _get_col_as_array(sheet, 1, 1)))

    value = _get_excel_metadata_value(sheet, TISSUE_SAMPLING_PERIOD_UUID)
    if value is None:
        raise NotImplementedError("Tissue Sampling Period should not be None here")
    sampling_period = int(round(1 / float(value), 6) * MICRO_TO_BASE_CONVERSION)

    interpolation_value_str = _get_excel_metadata_value(sheet, INTERPOLATION_VALUE_UUID)
    interpolation_value = (
        float(sampling_period)
        if interpolation_value_str is None
        else float(interpolation_value_str) * MICRO_TO_BASE_CONVERSION
    )

    begin_recording = _get_excel_metadata_value(sheet, UTC_BEGINNING_RECORDING_UUID)
    begin_recording = datetime.datetime.strptime(begin_recording, "%Y-%m-%d %H:%M:%S")  # type: ignore
    well_name = _get_excel_metadata_value(sheet, WELL_NAME_UUID)

    attrs = {
        FILE_FORMAT_VERSION_METADATA_KEY: NOT_APPLICABLE_LABEL,
        TISSUE_SENSOR_READINGS: raw_tissue_reading,
        REFERENCE_SENSOR_READINGS: np.zeros(raw_tissue_reading.shape),
        str(INTERPOLATION_VALUE_UUID): interpolation_value,
        str(TISSUE_SAMPLING_PERIOD_UUID): sampling_period,
        str(UTC_BEGINNING_RECORDING_UUID): begin_recording,
        str(MANTARRAY_SERIAL_NUMBER_UUID): _get_excel_metadata_value(sheet, MANTARRAY_SERIAL_NUMBER_UUID),
        str(PLATE_BARCODE_UUID): _get_excel_metadata_value(sheet, PLATE_BARCODE_UUID),
        str(WELL_NAME_UUID): well_name,
        str(DATA_TYPE_UUID): _get_data_type(_get_excel_metadata_value(sheet, DATA_TYPE_UUID)),
    }

    return attrs


def _get_data_type(data_type_meta: Optional[str]):
    default = "Calcium"

    if data_type_meta is None:
        return default
    if "calcium" in data_type_meta.lower():
        return "Calcium"
    if "voltage" in data_type_meta.lower():
        return "Voltage"

    return default
