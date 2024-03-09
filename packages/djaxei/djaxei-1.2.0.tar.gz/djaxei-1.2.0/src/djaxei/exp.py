import os
from collections import OrderedDict
from collections.abc import Iterable
from tempfile import NamedTemporaryFile

from django.contrib.admin.utils import NestedObjects
from django.db import router, models
from django.db.models.query import QuerySet
from openpyxl import Workbook


class AbstractMoDem:

    def __init__(self, model, rules) -> None:
        """The base exporter class.

        :param model:
        :param rules:
        """
        self.rules = rules
        if isinstance(model, str):  # we assume 'app.modelname' format
            self.model = model.strip().lower()
        elif isinstance(model, models.Model):
            self.model = model._meta.label_lower

    def modulate(self, obj: models.Model, context):
        pass

    def demodulate(self, obj, context):
        pass


class SimpleMoDem:
    """A simple MoDem class accepting e list of either directly fieldnames or (fieldname, FieldMoDem).

    :param model:
    :param rules:
    """
    def modulate(self, obj: models.Model, context):
        row = []
        for x in self.rules:
            if isinstance(x, str):
                row.append(getattr(obj, x))
            else:  # we consider FieldMoDem
                row.append(x[1]().modulate(getattr(obj, x[0])))
        return row



class ExcelMoDem(AbstractMoDem):

    def modulate(self, obj: models.Model, context):
        fields = obj._meta.label_lower
        row = []
        for x in fields:
            if isinstance(x, str):
                row.append(getattr(obj, x))
            else:
                row.append(x[1](getattr(obj, x[0])))
        return row


class Exporter:
    def __init__(self, root, modems: list):
        """Create an exporter.

        :param root: Can be a Django object, or a queryset
        :param rules: A dictionary 'appname.modelname': list of field names or tuple(fieldname, stringify function)
        """
        if isinstance(root, QuerySet):
            self.roots = root
            self.using = router.db_for_write(root.first()._meta.model)
        elif isinstance(root, Iterable):
            self.roots = root
            self.using = router.db_for_write(list(root)[0]._meta.model)
        else:
            self.roots = [root]
            self.using = router.db_for_write(root._meta.model)
        self.modems = modems

        # {}
        # for model_ref, field_refs in rules.items():
        #     if isinstance(model_ref, str):
        #         model_ref = model_ref.lower()
        #     else:
        #         model_ref = model_ref._meta.label_lower
        #     self.mode[model_ref] = field_refs

    def get_modem_for_object(self, obj):
        ret = None
        for modem in self.modems:
            if modem.model_label == obj._meta.label_lower:
                ret = modem
                break
        return ret


    def xls_export(self, target):
        # lmodels = {}

        wb = Workbook(write_only=True)
        sheets = OrderedDict()

        for modem in self.modems:
            header = modem.get_header()
            sheets[modem.model_label] = [header, ]
            wb.create_sheet(modem.model_label)

        collector = NestedObjects(using=self.using)
        collector.collect(self.roots)

        def callback_generator(exporter):
            exporter = exporter
            def callback(obj):
                modem = exporter.get_modem_for_object(obj)
                sheets[obj._meta.label_lower].append(modem.modulate(obj))

                # fields = lmodels.get(obj._meta.label_lower, None)
                # if fields:
                #     row = []
                #     for x in fields:
                #         if isinstance(x, str):
                #             row.append(getattr(obj, x))
                #         else:
                #             row.append(x[1](getattr(obj, x[0])))
                #     sheets[obj._meta.label_lower].append(row)
            return callback

        collector.nested(callback_generator(self))

        for sheet_name, rows in sheets.items():
            for row in rows:
                wb[sheet_name].append(row)

        wb.save(target)
        #
        # workbook = None
        # workbookfile = None
        # try:
        #     sheets = OrderedDict()
        #
        #     lmodels = {}
        #     for k, v in _models.items():
        #         lname = k.lower()
        #         model_name = lname.rsplit('.')[1]
        #         lmodels[lname] = v
        #         sheets[model_name] = [v, ]
        #
        #     if root:
        #         root_qs = root._meta.model.objects.filter(pk=root.pk)
        #
        #     using = router.db_for_write(root_qs.first()._meta.model)
        #     collector = NestedObjects(using=using)
        #     collector.collect(root_qs)
        #
        #     def callback(obj):
        #         fields = lmodels.get(obj._meta.label_lower, None)
        #         if fields:
        #             sheets[obj._meta.model_name].append([getattr(obj, x) for x in fields])
        #
        #     collector.nested(callback)
        #
        #     Workbook = get_workbook_impl()
        #     workbookfile = self.dest or NamedTemporaryFile(dir=self.tmpdir, suffix=Workbook._preferred_suffix, delete=False)
        #     Workbook(workbookfile).write_data(sheets)
        #
        #     return workbookfile.name
        #
        # except Exception as e:
        #     if workbook:
        #         if not workbookfile.closed:
        #             workbookfile.close()
        #         if os.path.exists(workbookfile.name):
        #             os.remove(workbookfile.name)
        #     raise e