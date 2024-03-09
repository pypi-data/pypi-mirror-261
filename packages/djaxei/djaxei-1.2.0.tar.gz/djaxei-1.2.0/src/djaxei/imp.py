from openpyxl.reader.excel import load_workbook


class Importer:
    def __init__(self, modems: list):
        self.modems = modems

    def xls_import(self, file):
        wb = load_workbook(file, read_only=True, data_only=True)

        remappings = {
            sheet_name: {} for sheet_name in wb.sheetnames
        }

        for modem in self.modems:
            modem.get_importer(remappings, wb).run()

#
# class Importer_:
#     def __init__(self, tmpdir=None, **kwargs):
#         self.tmpdir = tmpdir
#
#     def xls_import(self, file, models_dict, *args, **kwargs):
#         wb = load_workbook(file, data_only=True)
#
#         for ws_name, callback in models_dict.items():
#             ws = wb[ws_name]
#             for rownum, row in enumerate(ws.iter_rows()):
#                 row = [cell.value for cell in row]
#                 if rownum == 0:
#                     header = row
#                 else:
#                     try:
#                         data_dict = OrderedDict(zip(header, row))
#                         data_dict = {k: v for k, v in data_dict.items() if k != None}
#                         callback(data_dict)
#                     except Exception as e:
#                         raise ImportException(e, worksheet=ws_name)
#         wb.close()

