import openpyxl


def excel_open(s_filepath):
    book = openpyxl.load_workbook(s_filepath + '.xlsx')
    return book


def excel_sheet(book):
    sheet = book['sides']
    return sheet


def excel_sheets_rows(sheet, fl_new):
    row = 0
    i = 1
    while i != 0:
        if sheet.cell(i, 1).value == None:
            i = 0
        else:
            row += 1
            i += 1
    if fl_new == 1:
        row += 1
    return row


def f_word(sheet, n_rn, n_cl):
    return sheet.cell(row=n_rn, column=n_cl).value


def s_word(sheet, n_rn, n_cl):
    word = sheet.cell(row=n_rn, column=n_cl + 2).value
    return word


def s_pass(sheet, n_rn, n_cl):
    spass = sheet.cell(row=n_rn, column=n_cl + 1).value
    return spass


def input_xls(s_str, excel_path, n_rn):
    wbook = excel_open(excel_path)
    wsheet = wbook['sides']
    n_cl = 2
    for word in s_str.split():
        n_cl += 1
        wsheet.cell(row=n_rn, column=n_cl).value = word
    wsheet.cell(row=n_rn, column=1).value = n_cl-2
    wsheet.cell(row=n_rn, column=2).value = 'metamask' + str(n_rn) + 'pass'
    wbook.save(excel_path + '.xlsx')