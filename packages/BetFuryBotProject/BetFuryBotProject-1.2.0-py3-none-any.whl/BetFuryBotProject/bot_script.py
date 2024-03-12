import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
import web3
import time
import requests
import openpyxl
import re

def urljoin(*args):
    return "\\".join(map(lambda x: str(x).rstrip('\\'), args))


def regexp(exp, mes, arg=''):
    while not re.match(exp, arg):
        arg = input(mes)
    return arg


def urlcheck(surl):
    r = requests.get(surl)
    return r.status_code


def network_param(s_network_name, s_param_name):
    match s_network_name.split():
        case ["bsc"]:
            unit_to_bsc_params = {
                'name_network': 'Binance Smart Chain BEP-20',
                'url_rpc': 'https://bsc-dataseed.binance.org/',
                'chain_id': '56',
                'currency_symbol': 'BNB',
                'block_explorer_url': 'https://bscscan.com'
            }
        case ["polygon"]:
            unit_to_bsc_params = {
                'name_network': 'Polygon Mainnet',
                'url_rpc': 'https://rpc-mainnet.matic.network',
                'chain_id': '137',
                'currency_symbol': 'MATIC',
                'block_explorer_url': 'https://polygonscan.com'
            }
        case ["cronos"]:
            unit_to_bsc_params = {
                'name_network': 'Cronos Mainnet',
                'url_rpc': 'https://rpc.vvs.finance',
                'chain_id': '25',
                'currency_symbol': 'cro',
                'block_explorer_url': 'https://cronos.crypto.org/explorer/'
            }
        case ["optimism"]:
            unit_to_bsc_params = {
                'name_network': 'Optimism Mainnet',
                'url_rpc': 'https://mainnet.optimism.io',
                'chain_id': '10',
                'currency_symbol': 'ETH',
                'block_explorer_url': 'https://optimistic.etherscan.io'
            }
        case ["arbitrum"]:
            unit_to_bsc_params = {
                'name_network': 'Arbitrum Mainnet',
                'url_rpc': 'https://arb1.arbitrum.io/rpc',
                'chain_id': '42161',
                'currency_symbol': 'ETH',
                'block_explorer_url': 'https://arbiscan.io'
            }
        case ["heco"]:
            unit_to_bsc_params = {
                'name_network': 'Huobi ECO Chain Mainnet',
                'url_rpc': 'https://http-mainnet-node.huobichain.com',
                'chain_id': '128',
                'currency_symbol': 'HT',
                'block_explorer_url': 'https://hecoinfo.com/'
            }
        case ["moonriver"]:
            unit_to_bsc_params = {
                'name_network': 'Moonriver',
                'url_rpc': 'https://rpc.api.moonriver.moonbeam.network',
                'chain_id': '1285',
                'currency_symbol': 'MOVR',
                'block_explorer_url': 'https://moonriver.moonscan.io'
            }
        case ["okex"]:
            unit_to_bsc_params = {
                'name_network': 'OKEXChain',
                'url_rpc': 'https://exchainrpc.okex.org',
                'chain_id': '66',
                'currency_symbol': 'OKT',
                'block_explorer_url': 'https://www.oklink.com/okexchain/'
            }
        case ["aurora"]:
            unit_to_bsc_params = {
                'name_network': 'Aurora',
                'url_rpc': 'https://mainnet.aurora.dev/',
                'chain_id': '1313161554',
                'currency_symbol': 'ETH',
                'block_explorer_url': 'https://explorer.mainnet.aurora.dev/'
            }
        case ["avalanche"]:
            unit_to_bsc_params = {
                'name_network': 'Avalanche Network',
                'url_rpc': 'https://api.avax.network/ext/bc/C/rpc',
                'chain_id': '43114',
                'currency_symbol': 'AVAX',
                'block_explorer_url': 'https://snowtrace.io/'
            }
        case ["boba"]:
            unit_to_bsc_params = {
                'name_network': 'BOBA L2',
                'url_rpc': 'https://mainnet.boba.network',
                'chain_id': '288',
                'currency_symbol': 'ETH',
                'block_explorer_url': 'https://blockexplorer.boba.network'
            }
    r_param_value = unit_to_bsc_params.get(s_param_name, 0)
    return r_param_value


def meta_reg(driver, sheet, spass, rn, delay):
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div/div/button')))
    except TimeoutException:
        print('Hello page not loaded...')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div/div[2]/div/div[2]/div[1]/button')))
    except TimeoutException:
        print('Choice page not loaded...')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div/div/div[5]/div[1]/footer/button[1]')))
    except TimeoutException:
        print('Create page not loaded...')
    else:
        element.click()
    n_word = int(f_word(sheet, rn, 0))
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="create-new-vault__terms-checkbox"]')))
    except TimeoutException:
        print('Create page not loaded...')
    else:
        for i in range(n_word):
            driver.find_element(
                By.XPATH, '//*[@id="import-srp__srp-word-' + str(i) + '"]')\
                .send_keys(s_word(sheet, rn, i))
        driver.find_element(
            By.XPATH, '//*[@id="password"]')\
            .send_keys(spass)
        driver.find_element(
            By.XPATH, '//*[@id="confirm-password"]')\
            .send_keys(spass)
        element.click()
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div/div[2]/form/button')\
            .click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div/button')))
    except TimeoutException:
        print('Congratulation page not loaded...')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="popover-content"]/div/div/section/header/div/button')))
    except TimeoutException:
        print('News page not loaded...')
    else:
        element.click()


def meta_network_reg(driver, delay, s_network):
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[1]/div/div[2]/div[1]/div/span')))
    except TimeoutException:
        print('Network dropdown list not loaded... button')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/button')))
    except TimeoutException:
        print('Network add menu not loaded... button')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[2]/div[1]/label/input')))
    except TimeoutException:
        print('Network add page not loaded... button')
    else:
        element.send_keys(network_param(s_network, 'name_network'))
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/label/input')\
            .send_keys(network_param(s_network, 'url_rpc'))
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[2]/div[3]/label/input')\
            .send_keys(network_param(s_network, 'chain_id'))
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[2]/div[5]/label/input')\
            .send_keys(network_param(s_network, 'block_explorer_url'))
        time.sleep(1.5)
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[2]/div[4]/label/input')\
            .send_keys(network_param(s_network, 'currency_symbol'))
        time.sleep(1.5)
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div[2]/div[2]/div/div[2]/div/div[3]/button[2]')\
            .click()


def meta_unl(driver, spass, delay):
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))
    except TimeoutException:
        print('Unlock page not loaded...')
    else:
        element.send_keys(spass)
        driver.find_element(
            By.XPATH, '//*[@id="app-content"]/div/div[3]/div/div/button')\
            .click()


def excel_open(s_filepath):
    book = openpyxl.load_workbook(s_filepath + '.xlsx')
    return book


def excel_sheet(book):
    sheet = book['sides']
    return sheet

def f_word(sheet, n_rn, n_cl):
    return sheet.cell_value(rowx=n_rn, colx=n_cl)


def s_word(sheet, n_rn, n_cl):
    word = sheet.cell_value(rowx=n_rn, colx=n_cl + 2)
    return word


def s_pass(sheet, n_rn, n_cl):
    spass = sheet.cell_value(rowx=n_rn, colx=n_cl + 1)
    return spass


def new_tabs_ch(driver):
    b = driver.current_window_handle
    for handle in driver.window_handles:
        if handle != b:
            driver.switch_to.window(handle)
            driver.close()
            driver.switch_to.window(b)


def meta_network_select(driver, delay, select_network):
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[1]/div/div[2]/div[1]/div/span')))
    except TimeoutException:
        print('Network dropdown list not loaded... button')
    else:
        element.click()
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="app-content"]/div/div[2]/div/div[2]/li[12]/span')))
    except TimeoutException:
        print('Network dropdown list not loaded... network')
    else:
        element.click()


def create_eth_wallet():
    from web3.auto import w3
    from mnemonic import Mnemonic

    mnemo = Mnemonic('english')
    words = mnemo.generate(strength=256)
    seed = mnemo.to_seed(words, passphrase="")
    entropy = mnemo.to_entropy(words)
    w3.eth.account.enable_unaudited_hdwallet_features()
    account = w3.eth.account.from_mnemonic(words)
    return (words)


def input_xls(s_str, excel_path, n_rn):
    wbook = excel_open(excel_path)
    wsheet = wbook['sides']
    n_cl = 2
    for word in s_str.split():
        n_cl += 1
        wsheet.cell(row=n_rn, column=n_cl).value = word
    wsheet.cell(row=n_rn, column=1).value = n_cl-2
    wsheet.cell(row=n_rn, column=2).value = 'metamask' + str(n_rn) + 'pass'
    wbook.save(excel_path + '.xlsx')


def run_chrome():
    root_dir = os.path.dirname(os.path.abspath(__file__))
    metamask_path = urljoin(root_dir, r'MetaMask\10.12.3_0.crx')
    excel_path = urljoin(root_dir, r'Sid_Phrases\sides')
    ch_dr_path = urljoin(root_dir, r'chromedriver_win32\chromedriver.exe')
    profiles_path = urljoin(root_dir, r'Data_profiles\profile_')
    metamask_extension_url = 'chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#initialize/welcome'

    book = excel_open(excel_path)
    sheet = excel_sheet(book)
    #print(str(sheet.max_row))
    #for row in range(1,sheet.max_row + 1):
    #    print(sheet.cell(row, 1).value)
    rn = sheet.max_row
    #input()
    b_add_wallets = regexp(r'^[yn]$', 'Do You want add wallets?(y/n) ')
    qty_wallets = 0
    if b_add_wallets == "y":
        qty_wallets = int(regexp(r'^\d+$', 'Write number wallets: '))
        for n_wallet in range(qty_wallets):
            words = create_eth_wallet()
            input_xls(words, excel_path, rn + n_wallet)
    book = excel_open(excel_path)
    sheet = excel_sheet(book)
    rn = sheet.max_row
    options = Options()
    options.add_argument('start-maximized')
    options.add_argument('user-data-dir='+profiles_path)
    options.add_extension(metamask_path)
    driver = webdriver.Chrome(executable_path=ch_dr_path, options=options)
    #ac = ActionChains(driver)
    delay = 10  # seconds
    driver.get(metamask_extension_url)
    time.sleep(1)
    new_tabs_ch(driver)
    #flurl = urlcheck('https://google.com')
    select_network = "bsc"
    spass = s_pass(sheet, rn, 0)
    if driver.current_url == metamask_extension_url:
        meta_reg(driver, sheet, spass, rn, delay)
        for s_network in ['bsc', 'polygon', 'cronos', 'optimism', 'arbitrum', 'heco', 'moonriver', 'okex', 'aurora', 'avalanche', 'boba']:
            meta_network_reg(driver, delay, s_network)
    else:
        meta_unl(driver, spass, delay)
    meta_network_select(driver, delay, select_network)
    input()
    #driver.close()


run_chrome()