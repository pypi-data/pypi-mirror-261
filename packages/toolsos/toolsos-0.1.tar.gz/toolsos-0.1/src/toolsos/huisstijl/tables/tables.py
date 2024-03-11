import functools
import json
import re
from typing import Any, Callable, Dict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

Fmt = list[list[dict[str, Any]]]
Mapping = Dict[str, Dict[str, str | int | bool]]

LOOKUP: dict[str, Callable] = {
    "font": Font,
    "fill": PatternFill,
    "alignment": Alignment,
    "border": Border,
    "protection": Protection,
    "side": Side,
    "border": Border,
}


def set_global_style(style: str) -> None:
    global STYLES
    if style == "old":
        from table_styles import STYLE_OLD

        STYLES = STYLE_OLD
    elif style == "new":
        from table_styles import STYLE_NEW

        STYLES = STYLE_NEW


# We are currently in the process of switching to the new `huisstijl`. Therefore the
# styling is stored in a json. After loading is treated as a constant
# STYLES = get_table_style_from_json()

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def df_to_array(df: pd.DataFrame) -> np.ndarray:
    """Turn dataframe into array that includes column names as the first row

    Args:
        df (pd.DataFrame): Dataframe to be turned into an array

    Returns:
        np.array: Array that includes the data and the column names as the first row
    """
    return np.vstack([df.columns, df.to_numpy()])


def get_fmt_table(arr: np.ndarray) -> Fmt:
    """Create nested list with dictionary inside that is the same size as the original
    dataframe including column names

    Args:
        df (pd.DataFrame): Dataframe to be written to excel

    Returns:
        Fmt: Return empty nest list that will later be used to store formatting info
    """
    fmt = []
    for _ in range(arr.shape[0] + 1):
        row: list = []
        for _ in range(arr.shape[1]):
            row.append({})
        fmt.append(row)

    return fmt


def update_format(fmt: Fmt, row_idx: int, col_idx: int, mapping: Mapping) -> Fmt:
    """Update the cell containing the formatting info

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idx (int): row index
        col_idx (int): column index
        mapping (Mapping): formatting info

    Returns:
        Fmt: nested list containing the updated formatting info
    """
    for fmt_type, args in mapping.items():
        cell = fmt[row_idx][col_idx].get(fmt_type)

        if not cell:
            fmt[row_idx][col_idx][fmt_type] = args
        else:
            fmt[row_idx][col_idx][fmt_type] = cell | args

    return fmt


def set_style_all(fmt: Fmt, mapping: Mapping) -> Fmt:
    """Set the formatting for all cells

    Args:
        fmt (Fmt): nested list containing the formatting info
        mapping (Mapping): formatting info

    Returns:
        Fmt: nested list containing the formatting info
    """
    for row_idx, row in enumerate(fmt):
        for col_idx, _ in enumerate(row):
            update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def set_style_row(
    fmt: Fmt,
    row_idxs: int | list,
    mapping: Mapping,
    exlude_col_ids: int | list | None = None,
) -> Fmt:
    """Set the formatting on a row

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idxs (int | list): The indices of the rows to be updated
        mapping (Mapping): formatting info
        exlude_col_ids (int | list | None, optional): Indices of the cols to be excluded
        when setting formatting for a row. Defaults to None.

    Returns:
        Fmt: nested list containing the formatting info
    """
    if isinstance(row_idxs, int):
        row_idxs = [row_idxs]

    if exlude_col_ids is not None:
        if isinstance(exlude_col_ids, int):
            exlude_col_ids = [exlude_col_ids]
    else:
        exlude_col_ids = []

    for row_idx in row_idxs:
        for col_idx, _ in enumerate(fmt[row_idx]):
            if col_idx not in exlude_col_ids:
                update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def set_style_col(
    fmt: Fmt,
    col_idxs: int | list,
    mapping: Mapping,
    exlude_row_ids: int | list | None = None,
) -> Fmt:
    """Set the formatting on a row

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idxs (int | list): The indices of the rows to be updated
        mapping (Mapping): formatting info
        exlude_row_ids (int | list | None, optional): Indices of the rows to be excluded
        when setting formatting for a col. Defaults to None.

    Returns:
        Fmt: nested list containing the formatting info
    """
    if isinstance(col_idxs, int):
        col_idxs = [col_idxs]

    if exlude_row_ids is not None:
        if isinstance(exlude_row_ids, int):
            exlude_row_ids = [exlude_row_ids]
    else:
        exlude_row_ids = []

    for col_idx in col_idxs:
        for row_idx, _ in enumerate(fmt):
            if row_idx not in exlude_row_ids:

                update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def excel_style(row: int, col: int) -> str:
    """Convert given row and column number to an Excel-style cell name."""
    result: list = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    return "".join(result) + str(row)


def get_cols_id_with_pattern(df: pd.DataFrame, pattern: str) -> list[int]:
    """Get columns indices from columns matching a regex pattern

    Args:
        df (pd.DataFrame): Input dataframe
        pattern (str): regex pattern to get columns indices when matching

    Returns:
        list[int]: list with column indices matching pattern
    """
    return [idx for idx, col in enumerate(df.columns) if re.findall(pattern, col)]


def get_string_cols_ids(df: pd.DataFrame) -> list[int]:
    """Get column indices of string columns

    Args:
        df (pd.DataFrame): Input dataframe

    Returns:
        list[int]: list with column indices of string columns
    """
    return [i for i, dtype in enumerate(df.dtypes) if dtype == "O"]


def get_numeric_col_ids(df: pd.DataFrame) -> list[int]:
    """Get column indices of numeric columns

    Args:
        df (pd.DataFrame): Input dataframe

    Returns:
        list[int]: list with column indices of numeric columns
    """
    num_cols = df.select_dtypes("number").columns
    return [i for i, col in enumerate(df.columns) if col in num_cols]


def cell_formatting(
    arr: np.ndarray,
    default_format: Mapping | None = None,
    blue_row_ids: int | list | None = None,
    light_blue_row_ids: int | list | None = None,
    light_blue_col_ids: int | list | None = None,
    left_align_ids: int | list | None = None,
    right_align_ids: int | list | None = None,
    perc_col_ids: int | list | None = None,
    perc_col_format: str | None = None,
    blue_border: bool | None = None,
    number_format: str | None = None,
):
    """Function to create the nested list with the shape of the input data (including columns)
    containing dictionaries with the formatting

    Args:
        arr (np.ndarray): array representing the data
        default_format (Mapping | None, optional): Default format applied to all cells. Defaults to None.
        blue_row_ids (int | list | None, optional): The ids of the rows to be colored blue. Defaults to None.
        light_blue_row_ids (int | list | None, optional): _description_. Defaults to None.
        light_blue_col_ids (int | list | None, optional): _description_. Defaults to None.
        left_align_ids (int | list | None, optional): _description_. Defaults to None.
        right_align_ids (int | list | None, optional): _description_. Defaults to None.
        perc_col_ids (int | list | None, optional): _description_. Defaults to None.
        perc_col_format (str | None, optional): _description_. Defaults to None.
        blue_border (bool | None, optional): _description_. Defaults to None.
        number_format (str | None, optional): _description_. Defaults to None.

    Returns:
        _type_: _description_
    """
    fmt = get_fmt_table(arr)

    if default_format:
        fmt = set_style_all(fmt, default_format)

    if number_format:
        fmt = set_style_all(fmt, {"number_format": {"format": number_format}})

    if blue_row_ids:
        fmt = set_style_row(fmt, blue_row_ids, STYLES["blue_white"])

    if light_blue_row_ids:
        fmt = set_style_row(fmt, light_blue_row_ids, STYLES["light_blue"])

    if light_blue_col_ids:
        fmt = set_style_col(fmt, light_blue_col_ids, STYLES["light_blue"], blue_row_ids)

    if left_align_ids:
        fmt = set_style_col(fmt, left_align_ids, STYLES["left_align"])

    if right_align_ids:
        fmt = set_style_col(fmt, right_align_ids, STYLES["right_align"])

    if perc_col_ids:
        if not perc_col_format:
            perc_col_format = "0.0%"
        fmt = set_style_col(
            fmt, perc_col_ids, {"number_format": {"format": perc_col_format}}
        )

    if blue_border:
        fmt = set_style_row(fmt, arr.shape[0] - 1, STYLES["blue_border_bottom"])

    return fmt


def write_worksheet(
    ws: Any,
    arr: np.ndarray,
    fmt: Fmt,
    title: str | None = None,
    col_filter: bool | None = None,
    autofit_columns: bool | None = None,
) -> None:
    """Writing data to worksheet. Used for writing values to cells and formatting the cells
    and

    Args:
        ws (Any): openpyxl worksheet
        arr (np.ndarray): array containing the input data
        fmt (Fmt): nested list containing dictionaries with the formatting info per cell
        title (str | None, optional): Title to be inserted above the table. Defaults to None.
        col_filter (bool | None, optional): Set column filter in excel. Defaults to None.
        autofit_columns (bool | None, optional): Automatically fit column width. Defaults to None.
    """
    for row_idx, row in enumerate(arr):
        for col_idx, _ in enumerate(row):
            value = arr[row_idx][col_idx]
            # Cell indices are not zero-indexed but one-indexed
            cell = ws.cell(row_idx + 1, col_idx + 1, value)
            # Get formatting for specific cell
            cell_fmt = fmt[row_idx][col_idx]
            for t, kwa in cell_fmt.items():
                # The api for setting different kind of formatting options is not
                # consistent therefore depeding on the formatting we want to set we have
                # to use a different strategy
                if t == "number_format":
                    # cell.number_format = "0.0"
                    setattr(cell, t, kwa["format"])
                elif t.startswith("border"):
                    # cell.border = Border(bottom=Side(color="00a0e6"))
                    type_, side = t.split("_")
                    side_spec = Side(**kwa)
                    setattr(cell, type_, LOOKUP[type_](**{side: side_spec}))
                else:
                    # cell.font = Font(color="B1D9F5", bold=True)
                    setattr(cell, t, LOOKUP[t](**kwa))

    if col_filter:
        filters = ws.auto_filter
        filters.ref = f"A1:{excel_style(len(fmt), len(fmt[0]))}"

    if autofit_columns:
        _autofit_columns(ws)

    if title:
        _insert_title(ws, title)


# def _set_column_width(ws: Any, column_widths: list) -> None:
#     for i, column_number in enumerate(range(ws.max_column)):
#         column_letter = get_column_letter(column_letter)
#         column_width = column_widths[i]
#         ws.column_dimensions[column_letter].width = column_width


def _autofit_columns(ws: Any) -> None:
    column_letters = tuple(
        get_column_letter(col_number + 1) for col_number in range(ws.max_column)
    )
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].auto_fit = True


def _insert_title(ws: Any, title: str) -> None:
    ws.insert_rows(0)
    cell = ws.cell(1, 1, title)
    for t, kwa in STYLES["title_bold"].items():
        setattr(cell, t, LOOKUP[t](**kwa))


def write_table(
    data: pd.DataFrame | dict[str, pd.DataFrame],
    file: str,
    header_row: int = 0,
    title: str | dict[str, str] | None = None,
    total_row: bool | None = None,
    total_col: bool | None = None,
    right_align_ids: list | None = None,
    right_align_pattern: str | None = None,
    right_align_numeric: bool | None = True,
    left_align_ids: list | None = None,
    left_align_pattern: str | None = None,
    left_align_string: bool | None = True,
    perc_ids: list | None = None,
    perc_pattern: str | None = None,
    perc_col_format: str | None = None,
    blue_border: bool | None = True,
    number_format: str = "0.0",
    autofit_columns: bool | None = False,
    col_filter: bool | None = False,
    style: str = "old",
):
    """_summary_

    Args:
        data (pd.DataFrame | dict[pd.DataFrame]): dataframe or dicts with dataframes
        name (str): name of excel file
        header_row (int): Set the number of rows to be dark blue (zero-indexed). Defaults to 0 (top row)
        title (str): Set the title above the table. In the case of multiple tables provide a dict in
        which te keys correspond to the sheet name. Defaults to none
        total_row (bool, optional): Color bottom row blue
        total_col (bool, optional): Color last column blue.
        right_align_ids (list, optional): The ids of the columns to right align. Defaults to None
        right_align_pattern (str, optional): Pattern of columns to right align. Defaults to None.
        right_align_numeric (bool, optional): Right align numeric columns. Defaults to True.
        left_align_ids (list, optional): The ids of the columns to left align. Defaults to None.
        left_align_pattern (str, optional): Pattern of columns to left align. Defaults to None.
        left_align_string (bool, optional): Left align string columns. Defaults to True.
        perc_ids (list, optional): The ids of the columns to format as percentage. Defaults to None.
        perc_pattern (str, optional): The pattern of columns to format as percentage. Defaults to None.
        perc_col_format (str, optional): The formatting string of percentage columns. Defaults to None.
        col_filter (bool, optional): Set filter on columns. Defaults to False.
    """

    wb = Workbook()
    # Empty sheet is created on Workbook creation
    del wb["Sheet"]

    set_global_style(style)

    if not isinstance(data, dict):
        data = {"Sheet1": data}

    for sheet_name, df in data.items():
        arr = df_to_array(df)

        blue_rows = []
        light_blue_rows = []
        light_blue_cols = []
        r_align_ids = []
        l_align_ids = []
        p_ids = []
        title_tbl = None

        if isinstance(header_row, int):
            blue_rows.extend(list(range(0, header_row + 1)))

        if title:
            if isinstance(title, str):
                title_tbl = title
            elif isinstance(title, dict):
                title_tbl = title.get(sheet_name)

        if right_align_ids:
            r_align_ids.extend(right_align_ids)

        if right_align_pattern:
            r_align_ids.extend(get_cols_id_with_pattern(df, right_align_pattern))

        if right_align_numeric:
            r_align_ids.extend(get_numeric_col_ids(df))

        if left_align_ids:
            r_align_ids.extend(left_align_ids)

        if left_align_pattern:
            l_align_ids.extend(get_cols_id_with_pattern(df, left_align_pattern))

        if left_align_string:
            l_align_ids.extend(get_string_cols_ids(df))

        if perc_ids:
            p_ids.extend(perc_ids)

        if perc_pattern:
            r_id = get_cols_id_with_pattern(df, perc_pattern)
            p_ids.extend(r_id)
            r_align_ids.extend(r_id)

        if total_row:
            light_blue_rows.append(arr.shape[0] - 1)

        if total_col:
            light_blue_cols.append(arr.shape[1] - 1)

        ws = wb.create_sheet(sheet_name)

        fmt = cell_formatting(
            arr=arr,
            default_format=STYLES["calibri"],
            blue_row_ids=blue_rows,
            light_blue_row_ids=light_blue_rows,
            light_blue_col_ids=light_blue_cols,
            left_align_ids=l_align_ids,
            right_align_ids=r_align_ids,
            perc_col_ids=p_ids,
            perc_col_format=perc_col_format,
            number_format=number_format,
            blue_border=blue_border,
        )

        write_worksheet(
            ws=ws,
            arr=arr,
            fmt=fmt,
            title=title_tbl,
            col_filter=col_filter,
            autofit_columns=autofit_columns,
        )

    wb.save(file)
